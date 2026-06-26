"""Generic Excel cell/range edit tool behavior and A1 cell-reference parsing."""
from __future__ import annotations

import json
import logging
import os
import re
from pathlib import Path
from typing import Any, List

from quotation.excel_io import ensure_writable
from system.workspace_paths import coerce_write_path

logger = logging.getLogger(__name__)


def _parse_cell_ref(ref: str) -> tuple[int, int] | None:
    """将 A1 形式转为 (row: 1-based, col: 1-based)，失败返回 None。"""
    m = re.match(r"^([A-Za-z]+)(\d+)$", (ref or "").strip())
    if not m:
        return None
    col_str, row_str = m.group(1).upper(), m.group(2)
    col = 0
    for c in col_str:
        col = col * 26 + (ord(c) - ord("A") + 1)
    try:
        row = int(row_str)
        return (row, col) if row >= 1 and col >= 1 else None
    except ValueError:
        return None


def edit_excel(
    file_path: str,
    edits: List[dict[str, Any]],
    sheet_name: str | None = None,
    output_path: str | None = None,
) -> dict[str, Any]:
    """
    【普适性】编辑任意 Excel：按单元格或区域写入。不依赖报价单列结构，任意 .xlsx/.xlsm 可用。

    edits 每项为以下之一：
    - {"cell": "A1", "value": 任意}：单格写入
    - {"range": "A1:B2", "values": [[v1,v2],[v3,v4]]}：区域按行写入

    Returns:
        {"success": bool, "result": str, "error": str | None, "output_path": str}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "output_path": ""}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "output_path": ""}
    out_p = Path(output_path) if output_path else path
    if output_path:
        out_p = Path(coerce_write_path(output_path, default_filename=path.name))

    if not edits or not isinstance(edits, list):
        return {"success": False, "result": "", "error": "请提供 edits 数组（每项含 cell+value 或 range+values）", "output_path": ""}

    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        applied = 0
        for op in edits:
            if op.get("cell") is not None and "value" in op:
                cell_ref = str(op["cell"]).strip()
                parsed = _parse_cell_ref(cell_ref)
                if not parsed:
                    continue
                row, col = parsed
                ws.cell(row=row, column=col, value=op["value"])
                applied += 1
            elif op.get("range") is not None and op.get("values") is not None:
                range_ref = str(op["range"]).strip()
                parts = range_ref.split(":")
                if len(parts) != 2:
                    continue
                start = _parse_cell_ref(parts[0])
                end = _parse_cell_ref(parts[1])
                if not start or not end:
                    continue
                row_s, col_s = start
                row_e, col_e = end
                vals = op["values"]
                if not isinstance(vals, list):
                    continue
                for ri, row_vals in enumerate(vals):
                    if not isinstance(row_vals, list):
                        continue
                    for ci, v in enumerate(row_vals):
                        r, c = row_s + ri, col_s + ci
                        if r > row_e or c > col_e:
                            break
                        ws.cell(row=r, column=c, value=v)
                applied += 1
        ensure_writable(out_p)
        wb.save(out_p)
        return {"success": True, "result": json.dumps({"applied_edits": applied, "output_path": str(out_p)}, ensure_ascii=False), "error": None, "output_path": str(out_p)}
    except Exception as e:
        return {"success": False, "result": "", "error": str(e), "output_path": ""}
