# 报价 Agent 工具：从原始报价单 Excel 提取「第2行到 Total Excluding PPN不含税总价 所在行之前」的报价数据，供 LLM 使用

from __future__ import annotations

import logging
import os
import re
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any, List, Tuple

from openpyxl.utils import get_column_letter

from quotation.fill_enrich import infer_default_satuan, resolve_quote_specification
from quotation.fill_row_guard import validate_and_fix_fill_rows
from quotation.inquiry_backfill import backfill_inquiry_columns_if_empty
from quotation.layout import LINGWEI_LAYOUT, VANTSING_LAYOUT, QuotationTemplateLayout
from quotation.excel_io import ensure_writable, save_workbook
from system.workspace_paths import coerce_write_path

logger = logging.getLogger(__name__)

# 表头关键词：交货日期列、报价日期标签（用于定位填写格）
DELIVERY_DATE_COL_KEYWORDS = [
    "交货日期",
    "Tanggal Pengiriman",
    "Delivery Date",
    "Tanggal Pengiriman Barang",
]
QUOTATION_DATE_LABEL_KEYWORDS = [
    "报价日期",
    "Tanggal Penawaran",
    "Quotation Date",
]

# 凌威模板：footer 中「报价日期」标签常在 J:K 合并格，数值应在 L:R 合并格（旧逻辑写 c+1 会落在 J:K 内）。
QUOTE_DATE_VALUE_MIN_COL = 12  # L
QUOTE_DATE_VALUE_MAX_COL = 18  # R


def _copy_cell_style(source_cell, dest_cell) -> None:
    """
    将 source_cell 的格式（边框、底纹、字体、对齐、数字格式）复制到 dest_cell，
    避免填表时新建单元格使用默认样式导致虚线边框或异常底纹（如绿色块）。
    """
    try:
        if getattr(source_cell, "has_style", False) and getattr(source_cell, "_style", None) is not None:
            dest_cell._style = copy(source_cell._style)
            return
    except (TypeError, AttributeError):
        pass
    try:
        if getattr(source_cell, "font", None) is not None:
            dest_cell.font = copy(source_cell.font)
        if getattr(source_cell, "border", None) is not None:
            dest_cell.border = copy(source_cell.border)
        if getattr(source_cell, "fill", None) is not None:
            dest_cell.fill = copy(source_cell.fill)
        if getattr(source_cell, "alignment", None) is not None:
            dest_cell.alignment = copy(source_cell.alignment)
        if getattr(source_cell, "number_format", None) is not None:
            dest_cell.number_format = source_cell.number_format
    except (TypeError, AttributeError):
        logger.debug("复制单元格样式时部分属性失败", exc_info=True)


def _use_excel_line_total_formulas(layout: QuotationTemplateLayout) -> bool:
    """VANTSING standard template: row/footer totals use Excel formulas for user edits."""
    return layout.template_id == "vantsing"


def _excel_col_letter(col_1based: int) -> str:
    return get_column_letter(col_1based)


def _row_line_total_formula(layout: QuotationTemplateLayout, row: int) -> str:
    price_col = _excel_col_letter(layout.unit_price_col)
    qty_col = _excel_col_letter(layout.quote_qty_col)
    return f"={price_col}{row}*{qty_col}{row}"


def _apply_footer_total_formulas(
    ws,
    layout: QuotationTemplateLayout,
    total_row_1based: int,
    data_start_row: int,
    last_data_row: int,
    freight_value: float,
) -> None:
    """Write dynamic SUM / PPN / grand-total formulas matching 空白标准报价单.xlsx."""
    total_col = _excel_col_letter(layout.totals_value_col)
    _set_cell_value_merged_safe(
        ws,
        row=total_row_1based,
        col=layout.totals_value_col,
        value=f"=SUM({total_col}{data_start_row}:{total_col}{last_data_row})",
    )
    _set_cell_value_merged_safe(
        ws,
        row=total_row_1based + 1,
        col=layout.totals_value_col,
        value=f"={total_col}{total_row_1based}*0.11",
    )
    _set_cell_value_merged_safe(
        ws,
        row=total_row_1based + 2,
        col=layout.totals_value_col,
        value=freight_value,
    )
    _set_cell_value_merged_safe(
        ws,
        row=total_row_1based + 3,
        col=layout.totals_value_col,
        value=f"=SUM({total_col}{total_row_1based}:{total_col}{total_row_1based + 2})",
    )


# (min_row, max_row, min_col, max_col)
_MergeRange = Tuple[int, int, int, int]


def _unmerge_ranges_touching_rows(ws, row_start: int, row_end: int) -> None:
    """Unmerge any merged range overlapping [row_start, row_end] (inclusive, 1-based)."""
    to_unmerge: list[_MergeRange] = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= row_end and merged_range.max_row >= row_start:
            to_unmerge.append(
                (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
            )
    for min_row, max_row, min_col, max_col in to_unmerge:
        ws.unmerge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def _unmerge_ranges_from_row(ws, from_row_1based: int) -> list[_MergeRange]:
    """Unmerge all ranges with min_row >= from_row_1based; return removed ranges for restore."""
    saved: list[_MergeRange] = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= from_row_1based:
            saved.append(
                (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
            )
    for min_row, max_row, min_col, max_col in saved:
        ws.unmerge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )
    return saved


def _apply_merge_ranges(ws, ranges: list[_MergeRange], *, row_offset: int = 0) -> None:
    for min_row, max_row, min_col, max_col in ranges:
        ws.merge_cells(
            start_row=min_row + row_offset,
            start_column=min_col,
            end_row=max_row + row_offset,
            end_column=max_col,
        )


def _insert_data_rows_before_total(
    ws,
    total_row_1based: int,
    insert_count: int,
    *,
    style_row_1based: int | None = None,
) -> int:
    """
    Insert blank data rows before the Total/footer block.

    Footer merges (e.g. VANTSING C18:M18) must be removed before insert_rows; otherwise
    openpyxl shifts them onto new data rows and quote columns become read-only MergedCells.
    Returns the new 1-based Total row index.
    """
    if insert_count <= 0:
        return total_row_1based

    style_row = style_row_1based if style_row_1based is not None else total_row_1based - 1
    footer_merges = _unmerge_ranges_from_row(ws, total_row_1based)

    ws.insert_rows(total_row_1based, insert_count)

    new_data_end = total_row_1based + insert_count - 1
    _unmerge_ranges_touching_rows(ws, total_row_1based, new_data_end)

    max_col = min(ws.max_column, 20)
    for new_row in range(total_row_1based, new_data_end + 1):
        for col in range(1, max_col + 1):
            _copy_cell_style(ws.cell(row=style_row, column=col), ws.cell(row=new_row, column=col))

    _apply_merge_ranges(ws, footer_merges, row_offset=insert_count)
    return total_row_1based + insert_count


def _ensure_data_rows_before_total(
    ws,
    layout: QuotationTemplateLayout,
    fill_items: list[dict[str, Any]],
    total_row_1based: int | None,
) -> int | None:
    """
    VANTSING 模板默认仅 10 条数据行（8–17）；超出时在 Total 行前插入行并复制末行样式。
    返回更新后的 total_row_1based。
    """
    if total_row_1based is None or not fill_items:
        return total_row_1based
    target_rows = [
        _to_int_or_none(it.get("row"))
        for it in fill_items
        if _to_int_or_none(it.get("row")) is not None
    ]
    if not target_rows:
        return total_row_1based
    max_target = max(r for r in target_rows if r is not None)
    last_data_row = total_row_1based - 1
    if max_target <= last_data_row:
        return total_row_1based
    insert_count = max_target - last_data_row
    new_total_row = _insert_data_rows_before_total(ws, total_row_1based, insert_count)
    logger.info(
        "Inserted %s data row(s) before Total (layout=%s, max_target=%s)",
        insert_count,
        layout.template_id,
        max_target,
    )
    return new_total_row


def _normalize_sheet_view(ws) -> None:
    """
    Normalize worksheet view for exported files.
    Some templates are saved in page-break preview mode, which users perceive
    as dashed lines and large tinted areas in Excel/WPS.
    Also clears conditional formatting and removes green-ish fills from the
    entire worksheet to eliminate template artifacts.
    """
    try:
        if hasattr(ws, "views") and hasattr(ws.views, "sheetView") and ws.views.sheetView:
            for sv in ws.views.sheetView:
                sv.view = "normal"
        if getattr(ws, "sheet_view", None) is not None:
            ws.sheet_view.view = "normal"
    except Exception:
        logger.debug("normalize sheet view failed", exc_info=True)

    try:
        ws.page_setup.horizontalCentered = False
        ws.page_setup.verticalCentered = False
        ws.sheet_properties.pageSetUpPr = None
    except Exception:
        logger.debug("clear page setup properties failed", exc_info=True)

    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        logger.debug("clear conditional formatting failed", exc_info=True)

    try:
        from openpyxl.styles import PatternFill
        no_fill = PatternFill(fill_type=None)
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                fill = getattr(cell, "fill", None)
                if fill is None:
                    continue
                fg = getattr(fill.fgColor, "rgb", None) if fill.fgColor else None
                bg = getattr(fill.bgColor, "rgb", None) if fill.bgColor else None
                if fg and isinstance(fg, str) and _is_green_like(fg):
                    cell.fill = no_fill
                elif bg and isinstance(bg, str) and _is_green_like(bg):
                    cell.fill = no_fill
    except Exception:
        logger.debug("clear green fills failed", exc_info=True)


def _is_green_like(rgb_hex: str) -> bool:
    """Return True if the RGB hex string looks green-ish (high G, low R/B)."""
    hex_str = rgb_hex.lstrip("#")
    if len(hex_str) == 8:
        hex_str = hex_str[2:]
    if len(hex_str) != 6:
        return False
    try:
        r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
    except ValueError:
        return False
    return g > 150 and g > r * 1.3 and g > b * 1.3

# 边界行标识（用户指定：报价数据列到该行为止）
TOTAL_ROW_MARKER = "Total Excluding PPN不含税总价"

# 询价列表头关键词（用于自动识别列）
NAME_COL_KEYWORDS = [
    "询价货物名称", "询价货物", "Nama Permintaan Barang", "Nama Permintaan", "nama permintaan",
]
SPEC_COL_KEYWORDS = [
    "询价规格型号", "询价规格", "Spesifikasi dan Model Permintaan Barang",
    "Spesifikasi Permintaan", "Spesifikasi",
]
QTY_COL_KEYWORDS = ["Jumlah", "数量", "jumlah", "Quantity", "QTY"]


def _cell_value(cell) -> str:
    v = getattr(cell, "value", None)
    if v is None:
        return ""
    return str(v).strip()


def _to_int_or_none(value: Any) -> int | None:
    """Best-effort int conversion; return None on invalid input."""
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _to_float_or_none(value: Any) -> float | None:
    """Best-effort float conversion; return None on invalid input."""
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _set_cell_value_merged_safe(ws, row: int, col: int, value: Any) -> None:
    """
    在处理合并单元格时安全地写入值：
    - 若目标单元格为普通单元格，直接写入；
    - 若为合并区域中的「非左上角」MergedCell，则改为写入该合并区域的左上角单元格，
      避免 openpyxl 抛出 "'MergedCell' object attribute 'value' is read-only"。
    """
    try:
        from openpyxl.cell.cell import MergedCell  # type: ignore
    except Exception:  # pragma: no cover - 仅防御性兜底
        ws.cell(row=row, column=col, value=value)
        return

    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    # 定位包含该 MergedCell 的合并区域，并将值写入左上角单元格
    try:
        for merged_range in getattr(ws, "merged_cells", []).ranges:  # type: ignore[attr-defined]
            if (merged_range.min_row <= row <= merged_range.max_row) and (
                merged_range.min_col <= col <= merged_range.max_col
            ):
                master = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                master.value = value
                return
    except Exception:
        # 回退到直接写入目标单元格，若再次触发异常则交由调用方处理
        pass
    cell.value = value


def _detect_quotation_layout(ws) -> QuotationTemplateLayout:
    """识别凌威旧模板 vs VANTSING 标准报价单。"""
    for r in range(1, 11):
        for c in range(1, 20):
            v = _cell_value(ws.cell(row=r, column=c))
            if not v:
                continue
            if "PENAWARAN HARGA" in v:
                return VANTSING_LAYOUT
            if c == 6 and ("Product number" in v or "产品编号" in v):
                return VANTSING_LAYOUT
    return LINGWEI_LAYOUT


def _find_inquiry_header(all_rows: List[List[str]], layout: QuotationTemplateLayout) -> tuple[int, int, int, int]:
    """
    在前 N 行表头中定位询价列，返回 (header_row_idx, name_col, spec_col, qty_col)（0-based 列索引）。
    VANTSING 模板列位置固定，表头在第 6–7 行；凌威模板仍按关键词扫描。
    """
    if layout.template_id == "vantsing":
        for idx in range(min(layout.header_scan_rows, len(all_rows))):
            nc = _find_col_by_header(all_rows[idx], NAME_COL_KEYWORDS)
            if nc >= 0:
                return idx, nc, layout.inquiry_spec_col - 1, layout.inquiry_qty_col - 1
        return -1, -1, -1, -1

    for idx, header_row in enumerate(all_rows[: layout.header_scan_rows]):
        nc = _find_col_by_header(header_row, NAME_COL_KEYWORDS)
        if nc >= 0:
            sc = _find_col_by_header(header_row, SPEC_COL_KEYWORDS)
            qc = _find_col_by_header(header_row, QTY_COL_KEYWORDS)
            return idx, nc, sc, qc
    return -1, -1, -1, -1


def _detect_layout_from_rows(all_rows: List[List[str]]) -> QuotationTemplateLayout:
    """从已读取的行数据识别模板类型。"""
    for row in all_rows[:10]:
        for c, cell in enumerate(row):
            s = (cell or "").strip()
            if not s:
                continue
            if "PENAWARAN HARGA" in s:
                return VANTSING_LAYOUT
            if c == 5 and ("Product number" in s or "产品编号" in s):
                return VANTSING_LAYOUT
    return LINGWEI_LAYOUT


def extract_quotation_data(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    从原始报价单 Excel 提取报价数据：从第 2 行起到「Total Excluding PPN不含税总价」所在行的上一行止。

    - 第 1 行视为表头。
    - 数据区：第 2 行 ～ 第一个包含 TOTAL_ROW_MARKER 的行的上一行。
    - 支持 .xlsx / .xlsm；.xls 需调用方先转为 xlsx 或在此用 pandas 扩展。

    Returns:
        {"success": bool, "result": str, "error": str | None, "rows_count": int}
        result 为 Markdown 表格或 JSON 文本，便于 LLM 阅读。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "result": "", "error": "仅支持 .xlsx / .xlsm；.xls 请先另存为 xlsx", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "result": "", "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        # 先收集所有行（read_only 下 iter_rows 只能遍历一次，先找边界行）
        all_rows: List[List[str]] = []
        total_row_1based: int | None = None

        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            cells = [_cell_value(c) for c in row]
            all_rows.append(cells)

            # 任意单元格包含边界标识即视为「总价行」，数据区在其上一行结束
            for c in cells:
                if TOTAL_ROW_MARKER in (c or ""):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            logger.debug("关闭 workbook 失败", exc_info=True)
        return {"success": False, "result": "", "error": str(e), "rows_count": 0}

    if not all_rows:
        return {"success": True, "result": "表为空，无报价数据。", "error": None, "rows_count": 0}

    # 表头：第 1 行（索引 0）
    header = all_rows[0]
    # 数据行：第 2 行到「Total Excluding PPN」的上一行（不包含总价行）
    if total_row_1based is not None and total_row_1based >= 2:
        data_rows = all_rows[1 : total_row_1based - 1]
    else:
        data_rows = all_rows[1:]

    if not data_rows:
        return {"success": True, "result": "未找到数据行（或仅含总价行）。", "error": None, "rows_count": 0}

    # 输出为 Markdown 表格，便于 LLM 理解
    def escape_md(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ")

    col_count = max(len(header), max(len(r) for r in data_rows), 1)
    header_padded = header + [""] * (col_count - len(header))
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    lines = ["| " + " | ".join(escape_md(h) for h in header_padded) + " |", sep]
    for r in data_rows:
        r_padded = r + [""] * (col_count - len(r))
        lines.append("| " + " | ".join(escape_md(c) for c in r_padded) + " |")

    result_text = (
        f"报价数据（第2行至「Total Excluding PPN不含税总价」上一行，共 {len(data_rows)} 行；"
        "其中 Qty=询价数量/采购数量，不是库存）：\n\n"
        + "\n".join(lines)
        + f"\n\n（共 {len(data_rows)} 行。回复用户时请列出上表**全部**数据行，勿只列部分。）"
    )
    return {
        "success": True,
        "result": result_text,
        "error": None,
        "rows_count": len(data_rows),
    }


def _find_col_by_header(header: List[str], keywords_list: List[str]) -> int:
    """按表头关键词查找列索引，返回 -1 表示未找到。"""
    for i, cell in enumerate(header):
        val = (str(cell or "").strip()).lower()
        for kw in keywords_list:
            if kw.lower() in val:
                return i
    return -1


def _find_delivery_date_column(ws, max_header_rows: int = 4, max_cols: int = 30) -> int:
    """
    在前几行表头中按关键词查找「交货日期」列，返回 1-based 列号，未找到返回 0。
    """
    for row_1based in range(1, max_header_rows + 1):
        row_cells: List[str] = []
        for c in range(1, max_cols + 1):
            try:
                v = ws.cell(row=row_1based, column=c).value
                row_cells.append(str(v).strip() if v is not None else "")
            except Exception:
                break
        col0 = _find_col_by_header(row_cells, DELIVERY_DATE_COL_KEYWORDS)
        if col0 >= 0:
            return col0 + 1
    return 0


def _find_quotation_date_cell(
    ws, total_row_1based: int, search_rows: int = 20, max_cols: int = 30
) -> Tuple[int, int] | None:
    """
    在合计行下方的 footer 区域查找「报价日期」标签所在行，返回应填写日期的单元格 (row, col) 1-based。

    - 凌威报价单：标签在 J:K 合并格时，旧约定「标签右侧一格」仍落在 J:K 内，错误。
      若同一行存在列号在 L–R（12–18）内的合并区域，则日期写入该区域左上角。
    - 其他模板：无 L–R 合并时，回退为标签所在单元格的右侧一列 (c+1)。
    """
    label_row: int | None = None
    label_col: int | None = None
    for r in range(total_row_1based + 4, total_row_1based + search_rows + 1):
        for c in range(1, max_cols + 1):
            try:
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                s = str(val).strip()
                if any(kw in s for kw in QUOTATION_DATE_LABEL_KEYWORDS):
                    label_row, label_col = r, c
                    break
            except Exception:
                continue
        if label_row is not None:
            break
    if label_row is None or label_col is None:
        return None

    try:
        candidates: list = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= label_row <= merged_range.max_row:
                if (
                    merged_range.min_col >= QUOTE_DATE_VALUE_MIN_COL
                    and merged_range.min_col <= QUOTE_DATE_VALUE_MAX_COL
                ):
                    candidates.append(merged_range)
        single_row = [m for m in candidates if m.min_row == m.max_row == label_row]
        if single_row:
            prefer_l = [m for m in single_row if m.min_col == QUOTE_DATE_VALUE_MIN_COL]
            chosen = prefer_l[0] if prefer_l else min(single_row, key=lambda m: m.min_col)
            return (label_row, chosen.min_col)
        if candidates:
            on_row = [m for m in candidates if m.min_row <= label_row <= m.max_row]
            if on_row:
                chosen = min(on_row, key=lambda m: (m.min_row, m.min_col))
                return (label_row, chosen.min_col)
    except Exception:
        logger.debug("_find_quotation_date_cell merge scan failed", exc_info=True)

    return (label_row, label_col + 1)


def _extract_inquiry_items_smart_fallback(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 200,
) -> dict[str, Any]:
    """
    普适解析 fallback：不依赖 TOTAL_ROW_MARKER，读取工作表前 max_rows 行，
    在前 3 行中按关键词识别名称/规格/数量列，构建 items。供 extract_inquiry_items 在主逻辑失败或无数据时调用。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "items": [], "error": "请安装 openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "items": [], "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "items": [], "error": "仅支持 .xlsx / .xlsm", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "items": [], "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        rows: List[List[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(c).strip() if c is not None else "" for c in (row or [])])
        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            logger.debug("关闭 workbook 失败", exc_info=True)
        return {"success": False, "items": [], "error": str(e), "rows_count": 0}

    if not rows or len(rows) < 2:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    name_col = spec_col = qty_col = -1
    header_row_idx = 0
    layout = _detect_layout_from_rows(rows)
    header_row_idx, name_col, spec_col, qty_col = _find_inquiry_header(rows, layout)

    if name_col < 0:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    if layout.template_id == "vantsing":
        data_start_idx = layout.data_start_row - 1
    else:
        data_start_idx = header_row_idx + 1
    data_rows = rows[data_start_idx:]
    items: List[dict] = []
    for i, row_cells in enumerate(data_rows):
        row_num = (layout.data_start_row if layout.template_id == "vantsing" else header_row_idx + 2) + i
        product_name = (row_cells[name_col] if name_col < len(row_cells) else "").strip()
        specification = (row_cells[spec_col] if spec_col >= 0 and spec_col < len(row_cells) else "").strip()
        keywords = f"{product_name} {specification}".strip() if specification else product_name
        if not keywords:
            continue
        qty_val = 0
        if qty_col >= 0 and qty_col < len(row_cells):
            try:
                v = row_cells[qty_col]
                if v is not None and str(v).strip():
                    qty_val = int(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                pass
        items.append({
            "row": row_num,
            "product_name": product_name,
            "specification": specification,
            "keywords": keywords,
            "qty": qty_val,
        })

    return {"success": True, "items": items, "error": None, "rows_count": len(items)}


def extract_inquiry_items(
    file_path: str,
    sheet_name: str | None = None,
    col_mapping: dict | None = None,
) -> dict[str, Any]:
    """
    提取「询价货物名称」「询价规格型号」两列，输出供库存查询 Agent 用的列表。

    - 复用 data 区域识别逻辑（第2行～Total Excluding PPN 上一行）
    - 列识别：按表头匹配，或通过 col_mapping 指定 {name_col: int, spec_col: int}
    - keywords = product_name + " " + specification（空规格则仅名称）

    Returns:
        {"success": bool, "items": [...], "error": str | None, "rows_count": int}
        items: [{"row": 1, "product_name": "", "specification": "", "keywords": ""}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "items": [], "error": "请安装 openpyxl: pip install openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "items": [], "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "items": [], "error": "仅支持 .xlsx / .xlsm", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "items": [], "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        all_rows: List[List[str]] = []
        total_row_1based: int | None = None

        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            cells = [_cell_value(c) for c in row]
            all_rows.append(cells)
            for c in cells:
                if TOTAL_ROW_MARKER in (c or ""):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            logger.debug("关闭 workbook 失败", exc_info=True)
        return {"success": False, "items": [], "error": str(e), "rows_count": 0}

    if not all_rows:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    # 确定表头行与列索引（部分模板首行为标题，次行为表头）
    if col_mapping:
        name_col = col_mapping.get("name_col", col_mapping.get("product_name_col", -1))
        spec_col = col_mapping.get("spec_col", col_mapping.get("specification_col", -1))
        qty_col = col_mapping.get("qty_col", col_mapping.get("quantity_col", -1))
        header_row_idx = 0
        layout = LINGWEI_LAYOUT
    else:
        layout = _detect_layout_from_rows(all_rows)
        header_row_idx, name_col, spec_col, qty_col = _find_inquiry_header(all_rows, layout)

    if name_col < 0:
        # Fallback：用普适解析（不依赖 Total Excluding PPN 与固定表头）再尝试识别列
        fallback = _extract_inquiry_items_smart_fallback(file_path, sheet_name)
        if fallback.get("items"):
            fallback["_fallback_used"] = True
            fallback["error"] = None
            return fallback
        return {"success": False, "items": [], "error": "未找到询价货物名称列，请检查表头或提供 col_mapping", "rows_count": 0}

    # 数据行从表头下一行起（VANTSING 固定第 8 行），到 Total Excluding PPN 上一行
    if layout.template_id == "vantsing":
        data_start = layout.data_start_row - 1
    else:
        data_start = header_row_idx + 1
    if total_row_1based is not None and total_row_1based >= 2:
        data_end = total_row_1based - 1
    else:
        data_end = len(all_rows)
    data_rows = all_rows[data_start:data_end]

    if not data_rows:
        fallback = _extract_inquiry_items_smart_fallback(file_path, sheet_name)
        if fallback.get("items"):
            fallback["_fallback_used"] = True
            return fallback
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    # spec_col 可为 -1，表示无规格列；qty_col 可为 -1，表示无数量列
    items: List[dict] = []
    for i, row_cells in enumerate(data_rows):
        row_num = (layout.data_start_row if layout.template_id == "vantsing" else data_start + 1) + i
        product_name = (row_cells[name_col] if name_col < len(row_cells) else "").strip()
        specification = (row_cells[spec_col] if spec_col >= 0 and spec_col < len(row_cells) else "").strip()
        keywords = f"{product_name} {specification}".strip() if specification else product_name
        if not keywords:
            continue
        # 需求数量 qty：用于库存比对
        qty_val = 0
        if qty_col >= 0 and qty_col < len(row_cells):
            try:
                v = row_cells[qty_col]
                if v is not None and str(v).strip():
                    qty_val = int(float(str(v).replace(",", "")))
            except (ValueError, TypeError) as e:
                logger.debug("解析数量失败 row=%s: %s", row_num, e)
        inquiry_unit = ""
        if layout.inquiry_unit_col:
            unit_col = layout.inquiry_unit_col - 1
            if 0 <= unit_col < len(row_cells):
                inquiry_unit = (row_cells[unit_col] or "").strip()
        items.append({
            "row": row_num,
            "product_name": product_name,
            "specification": specification,
            "keywords": keywords,
            "qty": qty_val,
            "inquiry_unit": inquiry_unit,
        })

    return {
        "success": True,
        "items": items,
        "error": None,
        "rows_count": len(items),
    }


# 案例报价单模板：表头第 2 行，数据从第 3 行起，A=序号 B=询价货物名称 C=规格 E=数量
INQUIRY_DATA_START_ROW = 3
INQUIRY_COL_SEQ = 1   # A
INQUIRY_COL_NAME = 2  # B
INQUIRY_COL_SPEC = 3  # C
INQUIRY_COL_QTY = 5   # E


def fill_template_with_inquiry_items(
    template_path: str,
    items: List[dict[str, Any]],
    output_path: str,
    sheet_name: str = "询价单",
    allow_insert_rows: bool = False,
) -> dict[str, Any]:
    """
    用「询价行」列表填写案例报价单模板，生成可被 extract_inquiry_items / Work 流程处理的 Excel。

    - 复制模板到 output_path，不修改原模板。
    - 从第 3 行起写入：A=序号(1-based)，B=product_name，C=specification，E=qty。
    - 默认不插行（allow_insert_rows=False），只在模板可用行内写入，确保文档样式稳定。
    - 如需扩展行数可启用 allow_insert_rows=True（复杂模板可能导致样式/结构异常）。

    Returns:
        {"success": bool, "output_path": str, "filled_count": int, "error": str | None, "capacity": int, "truncated_count": int}
    """
    import shutil
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "output_path": "", "filled_count": 0, "error": "请安装 openpyxl"}

    tpl = Path(template_path)
    if not tpl.is_absolute():
        tpl = Path(os.getcwd()) / tpl
    if not tpl.exists():
        return {"success": False, "output_path": "", "filled_count": 0, "error": f"模板不存在: {tpl}"}
    out_p = Path(
        coerce_write_path(output_path, default_filename=f"{tpl.stem}_inquiry{tpl.suffix}")
    )
    try:
        shutil.copy2(tpl, out_p)
        ensure_writable(out_p)
    except Exception as e:
        return {"success": False, "output_path": "", "filled_count": 0, "error": str(e)}

    items = [x for x in items if isinstance(x, dict) and (x.get("product_name") or x.get("name"))]
    if not items:
        return {
            "success": True,
            "output_path": str(out_p),
            "filled_count": 0,
            "error": None,
            "capacity": 0,
            "truncated_count": 0,
        }

    try:
        wb = openpyxl.load_workbook(out_p)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        total_row_1based = None
        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            for cell in row:
                if TOTAL_ROW_MARKER in _cell_value(cell):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break
        if total_row_1based is None:
            total_row_1based = ws.max_row + 1
        layout = _detect_quotation_layout(ws)
        data_start = layout.data_start_row
        available = max(0, total_row_1based - data_start)
        truncated_count = 0
        if len(items) > available:
            if allow_insert_rows:
                insert_count = len(items) - available
                total_row_1based = _insert_data_rows_before_total(
                    ws, total_row_1based, insert_count
                )
            else:
                truncated_count = len(items) - available
                items = items[:available]
        filled = 0
        for i, it in enumerate(items):
            row_num = data_start + i
            name = (it.get("product_name") or it.get("name") or "").strip()
            spec = (it.get("specification") or it.get("spec") or "").strip()
            qty = _to_int_or_none(it.get("qty", 0))
            if qty is None:
                qty = 0
            ws.cell(row=row_num, column=layout.inquiry_seq_col, value=i + 1)
            ws.cell(row=row_num, column=layout.inquiry_name_col, value=name)
            ws.cell(row=row_num, column=layout.inquiry_spec_col, value=spec)
            unit = (
                (it.get("inquiry_unit") or it.get("satuan") or "").strip()
                or infer_default_satuan(name, spec, name)
            )
            if unit and layout.inquiry_unit_col:
                ws.cell(row=row_num, column=layout.inquiry_unit_col, value=unit)
            ws.cell(row=row_num, column=layout.inquiry_qty_col, value=max(0, qty))
            filled += 1
        _normalize_sheet_view(ws)
        save_workbook(wb, out_p)
        return {
            "success": True,
            "output_path": str(out_p),
            "filled_count": filled,
            "error": None,
            "capacity": available,
            "truncated_count": truncated_count,
        }
    except Exception as e:
        return {
            "success": False,
            "output_path": "",
            "filled_count": 0,
            "error": str(e),
            "capacity": 0,
            "truncated_count": 0,
        }


def get_template_inquiry_capacity(
    template_path: str,
    sheet_name: str = "询价单",
) -> dict[str, Any]:
    """
    读取询价模板在不插行模式下的可填写容量（可写行数）。
    capacity = Total 行号 - INQUIRY_DATA_START_ROW
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "capacity": 0, "error": "请安装 openpyxl"}

    path = Path(template_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "capacity": 0, "error": f"模板不存在: {path}"}

    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        total_row_1based = None
        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            for cell in row:
                if TOTAL_ROW_MARKER in _cell_value(cell):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break
        if total_row_1based is None:
            total_row_1based = ws.max_row + 1
        capacity = max(0, total_row_1based - INQUIRY_DATA_START_ROW)
        wb.close()
        return {"success": True, "capacity": capacity, "error": None}
    except Exception as e:
        return {"success": False, "capacity": 0, "error": str(e)}


# ---------------------------------------------------------------------------
# 普适性 Excel 工具（不依赖报价单结构，任意 Excel 可用）
# ---------------------------------------------------------------------------

def parse_excel_smart(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 500,
) -> dict[str, Any]:
    """
    【普适性】智能解析任意 Excel：自动读取指定工作表的所有单元格（或前 max_rows 行），
    零硬编码列/行，适合多表头、合并单元格、不规则布局。返回 Markdown 表格便于 LLM 理解。

    Returns:
        {"success": bool, "result": str, "error": str | None, "sheet_name": str, "rows_read": int}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "sheet_name": "", "rows_read": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "sheet_name": "", "rows_read": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "result": "", "error": "仅支持 .xlsx / .xlsm", "sheet_name": "", "rows_read": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "result": "", "error": f"打开 Excel 失败: {e}", "sheet_name": "", "rows_read": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            used_sheet = sheet_name
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
            used_sheet = ws.title

        rows: List[List[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(c).strip() if c is not None else "" for c in (row or [])])
        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"success": False, "result": "", "error": str(e), "sheet_name": used_sheet if "used_sheet" in dir() else "", "rows_read": 0}

    if not rows:
        return {"success": True, "result": "表为空或无数据。", "error": None, "sheet_name": used_sheet, "rows_read": 0}

    def escape_md(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ")

    col_count = max(len(r) for r in rows)
    lines = []
    for i, r in enumerate(rows):
        r_padded = list(r) + [""] * (col_count - len(r))
        lines.append("| " + " | ".join(escape_md(str(c)) for c in r_padded) + " |")
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    header_block = (
        f"工作表「{used_sheet}」共 {len(rows)} 行（普适解析，未限定列）：\n\n"
        + "| " + " | ".join(escape_md(str(i + 1)) for i in range(col_count)) + " |\n"
        + sep + "\n"
    )
    # 若结果过长，在完整行边界处截断，避免下游在行中间截断导致模型在单元格填「数据被截断」
    _max_result_chars = 40_000
    if len(header_block) + sum(len(ln) + 1 for ln in lines) + 200 > _max_result_chars:
        n_show = 0
        acc = len(header_block) + 200
        for ln in lines:
            if acc + len(ln) + 1 > _max_result_chars:
                break
            acc += len(ln) + 1
            n_show += 1
        lines_show = lines[:n_show] if n_show else lines[:1]
        result_text = (
            header_block + "\n".join(lines_show)
            + f"\n\n（因长度限制仅展示前 {len(lines_show)} 行，共 {len(rows)} 行。回复时按上表逐行照抄，勿在单元格内填「数据被截断」。）"
        )
    else:
        result_text = (
            header_block + "\n".join(lines)
            + f"\n\n（共 {len(rows)} 行。回复时**必须按上表逐行照抄**，不得只列部分、不得将同一行重复多遍凑数、不得自行编造行、勿在单元格内填「数据被截断」。）"
        )
    return {"success": True, "result": result_text, "error": None, "sheet_name": used_sheet, "rows_read": len(rows)}


# 凌威报价单回填列（1-based）
COL_PRODUCT_NO = 7   # G 产品编号
COL_QUOTE_NAME = 8   # H 报价名称
COL_QUOTE_SPEC = 10  # J 报价产品规格
COL_QTY_OUT = 12     # L 数量
COL_UNIT_PRICE = 14  # N 单价
COL_TOTAL = 15       # O 总价

# 4 个价格计算行（在 Total Excluding PPN 所在行及其后 3 行），金额写入列与 COL_TOTAL 一致
TOTALS_VALUE_COL = COL_TOTAL  # O 列

def fill_quotation(
    file_path: str,
    fill_items: list[dict[str, Any]],
    sheet_name: str | None = None,
    output_path: str | None = None,
    freight: float = 0.0,
    quotation_date: str | None = None,
    delivery_date: str | None = None,
) -> dict[str, Any]:
    """
    将匹配到的产品信息回填到报价单 Excel。
    每行保证填写的列：G=产品编号, H=报价名称, J=报价产品规格(无 specification 时用 quote_name), L=数量, N=单价, O=总价；
    未匹配项写「无货」；写完后更新底部 4 个价格计算行；并按表头填写「交货日期」「报价日期」。

    推荐入参：fill_items 来自「规范行」导出的 fill_items（见 canonical_lines.fill_items_from_canonical_lines），
    以保证 Excel 列 J（报价产品规）与草稿一致。fill_items 每项需含 row, code, quote_name, unit_price, qty, specification；
    写入列 J 时使用传入的 specification（即应由调用方传入规范行中的 quote_spec 或 specification）。

    Args:
        file_path: 原始报价单路径
        fill_items: 每项含 row, code, quote_name, unit_price, qty, specification；code="无货" 表示未匹配
        sheet_name: 工作表名，默认第一个
        output_path: 输出路径，默认覆盖原文件（建议调用方传副本路径）
        freight: 运费，默认 0
        quotation_date: 报价日期，默认当天 YYYY/MM/DD
        delivery_date: 交货日期（每行同一值），默认当天 YYYY/MM/DD

    Returns:
        {"success": bool, "output_path": str, "filled_count": int, "error": str | None}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "output_path": "", "filled_count": 0, "error": "请安装 openpyxl"}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "output_path": "", "filled_count": 0, "error": f"文件不存在: {path}"}
    out_p = Path(output_path) if output_path else path
    if output_path:
        out_p = Path(coerce_write_path(output_path, default_filename=path.name))
    load_path = path
    if out_p != path:
        import shutil
        if not out_p.exists():
            shutil.copy2(path, out_p)
        load_path = out_p
    ensure_writable(load_path)
    try:
        wb = openpyxl.load_workbook(load_path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        layout = _detect_quotation_layout(ws)
        fill_items = validate_and_fix_fill_rows(list(fill_items), layout)

        # 先定位「Total Excluding PPN」行，避免后续 iter_rows 与写表顺序导致样式错乱
        total_row_1based = None
        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            for cell in row:
                if TOTAL_ROW_MARKER in _cell_value(cell):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        total_row_1based = _ensure_data_rows_before_total(
            ws, layout, fill_items, total_row_1based
        )

        today_str = date.today().strftime("%Y/%m/%d")
        qdate_str = (quotation_date or today_str).strip() or today_str
        ddate_str = (delivery_date or today_str).strip() or today_str

        delivery_col = _find_delivery_date_column(ws)
        quotation_date_cell = _find_quotation_date_cell(ws, total_row_1based or 0) if total_row_1based else None
        if (
            layout.template_id == "vantsing"
            and layout.quote_date_value_col
            and quotation_date_cell
        ):
            qr, qc = quotation_date_cell
            quotation_date_cell = (qr, layout.quote_date_value_col)

        # Safe document-fill mode:
        # - only write values into target cells
        # - keep existing workbook styles as-is to avoid accidental style corruption
        #   (dashed borders, unexpected fills, etc.)
        freight_value = _to_float_or_none(freight)
        if freight_value is None:
            freight_value = 0.0

        filled = 0
        total_excluding_ppn = 0.0
        filled_data_rows: list[int] = []
        use_formulas = _use_excel_line_total_formulas(layout)
        for it in fill_items:
            row_num = _to_int_or_none(it.get("row"))
            if row_num is None or row_num <= 0:
                continue
            code = it.get("code")
            if code:
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.product_no_col, value=str(code))
                filled += 1
            if it.get("quote_name"):
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.quote_name_col, value=str(it["quote_name"]))
            up = _to_float_or_none(it.get("unit_price"))
            q = _to_int_or_none(it.get("qty"))
            if up is not None:
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.unit_price_col, value=up)
            if q is not None:
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.quote_qty_col, value=q)
            backfill_inquiry_columns_if_empty(
                ws,
                row_num,
                layout,
                it,
                satuan=str(it.get("satuan") or "") or None,
                qty=q,
            )
            spec_val = resolve_quote_specification(
                str(it.get("quote_name") or ""),
                str(it.get("specification") or ""),
                inquiry_spec=str(it.get("inquiry_spec") or ""),
                description_english=str(
                    it.get("description_english") or it.get("indonesian_name") or ""
                ),
            )
            _set_cell_value_merged_safe(
                ws,
                row=row_num,
                col=layout.quote_spec_col,
                value=spec_val if spec_val else None,
            )
            if layout.indonesian_name_col and it.get("indonesian_name"):
                _set_cell_value_merged_safe(
                    ws, row=row_num, col=layout.indonesian_name_col, value=str(it["indonesian_name"])
                )
            satuan_val = it.get("satuan")
            if not satuan_val and layout.inquiry_unit_col:
                satuan_val = _cell_value(ws.cell(row=row_num, column=layout.inquiry_unit_col))
            if not satuan_val:
                satuan_val = infer_default_satuan(
                    str(it.get("quote_name") or ""),
                    str(it.get("inquiry_spec") or it.get("specification") or ""),
                    str(it.get("product_name") or ""),
                )
            if layout.satuan_col and satuan_val:
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.satuan_col, value=str(satuan_val))
            if layout.brand_col and it.get("brand"):
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.brand_col, value=str(it["brand"]))
            remark = str(it.get("remark") or it.get("catatan") or "").strip()
            if layout.remark_col and remark:
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.remark_col, value=remark)
            if up is not None and q is not None and code and str(code) != "无货":
                row_total = up * q
                if use_formulas:
                    _set_cell_value_merged_safe(
                        ws,
                        row=row_num,
                        col=layout.total_col,
                        value=_row_line_total_formula(layout, row_num),
                    )
                else:
                    _set_cell_value_merged_safe(
                        ws, row=row_num, col=layout.total_col, value=round(row_total, 2)
                    )
                total_excluding_ppn += row_total
                filled_data_rows.append(row_num)
            elif up is not None and q is not None and (not code or str(code) == "无货"):
                _set_cell_value_merged_safe(ws, row=row_num, col=layout.total_col, value=0)
                filled_data_rows.append(row_num)
            if delivery_col:
                _set_cell_value_merged_safe(ws, row=row_num, col=delivery_col, value=ddate_str)

        ppn = round(total_excluding_ppn * 0.11, 2)
        total_including = round(total_excluding_ppn + ppn + freight_value, 2)
        if total_row_1based is not None:
            if use_formulas:
                last_data_row = max(filled_data_rows) if filled_data_rows else total_row_1based - 1
                _apply_footer_total_formulas(
                    ws,
                    layout,
                    total_row_1based,
                    layout.data_start_row,
                    last_data_row,
                    freight_value,
                )
            else:
                _set_cell_value_merged_safe(
                    ws,
                    row=total_row_1based,
                    col=layout.totals_value_col,
                    value=round(total_excluding_ppn, 2),
                )
                _set_cell_value_merged_safe(
                    ws,
                    row=total_row_1based + 1,
                    col=layout.totals_value_col,
                    value=ppn,
                )
                _set_cell_value_merged_safe(
                    ws,
                    row=total_row_1based + 2,
                    col=layout.totals_value_col,
                    value=freight_value,
                )
                _set_cell_value_merged_safe(
                    ws,
                    row=total_row_1based + 3,
                    col=layout.totals_value_col,
                    value=total_including,
                )
        if quotation_date_cell:
            qr, qc = quotation_date_cell
            _set_cell_value_merged_safe(ws, row=qr, col=qc, value=qdate_str)
        
        # 强制裁剪列：只保留 A-R (1-18列)，删除右侧所有列（包括绿色区域）
        try:
            MAX_COL = 18  # A-R
            current_max = ws.max_column
            logger.info(f"Excel 当前列数: {current_max}, 将裁剪到: {MAX_COL}")
            
            if current_max > MAX_COL:
                delete_count = current_max - MAX_COL
                ws.delete_cols(MAX_COL + 1, delete_count)
                logger.info(f"已删除 {delete_count} 列 (从第 {MAX_COL + 1} 列开始)")
                logger.info(f"裁剪后列数: {ws.max_column}")
            else:
                logger.info(f"列数 {current_max} <= {MAX_COL}, 无需裁剪")
        except Exception as e:
            logger.error(f"裁剪列失败: {e}", exc_info=True)
        
        _normalize_sheet_view(ws)
        save_workbook(wb, out_p)
        return {"success": True, "output_path": str(out_p), "filled_count": filled, "error": None}
    except ValueError as e:
        return {"success": False, "output_path": "", "filled_count": 0, "error": str(e)}
    except Exception as e:
        return {"success": False, "output_path": "", "filled_count": 0, "error": str(e)}


def get_quote_tools_openai_format() -> list[dict]:
    from quotation.tool_schema import get_quote_tools_openai_format as _get

    return _get()


def execute_quote_tool(name: str, arguments: dict[str, Any]) -> dict[str, Any]:
    from quotation.tool_adapter import execute_quote_tool as _execute

    return _execute(name, arguments)


def edit_excel(*args, **kwargs):
    from quotation.excel_edit import edit_excel as _edit

    return _edit(*args, **kwargs)
