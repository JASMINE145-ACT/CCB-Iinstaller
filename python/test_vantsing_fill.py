# -*- coding: utf-8 -*-
"""Manual test: VANTSING standard quotation fill on blank template."""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

import openpyxl
from quotation.quote_tools import (
    VANTSING_LAYOUT,
    _detect_quotation_layout,
    extract_inquiry_items,
    fill_quotation,
    fill_template_with_inquiry_items,
)

DATA = ROOT / "data"
OUT_DIR = Path(tempfile.mkdtemp(prefix="vantsing_fill_"))


def find_file(keyword: str) -> Path:
    matches = [p for p in DATA.glob("*.xlsx") if keyword in p.name]
    if not matches:
        raise FileNotFoundError(f"No xlsx matching {keyword!r} in {DATA}")
    return matches[0]


def main() -> int:
    blank = find_file("空白")
    filled_ref = find_file("已填")
    work = OUT_DIR / "test_vantsing_fill.xlsx"

    # Step 1: write inquiry rows into blank template
    inquiry_items = [
        {"product_name": "PE 给 pipe", "specification": "125mm/6m", "qty": 2},
        {"product_name": "PVC 排水 pipe", "specification": "DN110", "qty": 5},
        {"product_name": "直接 50", "specification": "", "qty": 10},
    ]
    tpl_result = fill_template_with_inquiry_items(str(blank), inquiry_items, str(work))
    print("fill_template_with_inquiry_items:", tpl_result)
    if not tpl_result.get("success"):
        return 1

    # Step 2: extract inquiry
    ext = extract_inquiry_items(str(work))
    print("extract_inquiry_items:", {k: ext[k] for k in ("success", "rows_count", "error")})
    for it in ext.get("items", []):
        print(" ", it)
    if not ext.get("success") or ext.get("rows_count", 0) != 3:
        print("FAIL: expected 3 inquiry items")
        return 1

    # Step 3: fill quotation (mock matched prices)
    fill_items = []
    mock_prices = [567858, 120000, 8500]
    for it, price in zip(ext["items"], mock_prices):
        fill_items.append({
            "row": it["row"],
            "code": f"TEST{it['row']}",
            "quote_name": f"测试报价-{it['product_name']}",
            "unit_price": price,
            "qty": it["qty"],
            "specification": it.get("specification") or "dn125",
        })

    fill_result = fill_quotation(str(work), fill_items, output_path=str(OUT_DIR / "test_vantsing_filled.xlsx"))
    print("fill_quotation:", fill_result)
    if not fill_result.get("success"):
        return 1
    work = Path(fill_result["output_path"])

    # Step 4: verify written cells
    wb = openpyxl.load_workbook(work, data_only=True)
    ws = wb.active
    layout = _detect_quotation_layout(wb.active)
    print("detected layout:", layout.template_id)
    assert layout.template_id == "vantsing", f"expected vantsing, got {layout.template_id}"

    errors = []
    for i, item in enumerate(fill_items):
        r = item["row"]
        f = ws.cell(r, layout.product_no_col).value
        g = ws.cell(r, layout.quote_name_col).value
        spec = ws.cell(r, layout.quote_spec_col).value
        k = ws.cell(r, layout.quote_qty_col).value
        m = ws.cell(r, layout.unit_price_col).value
        n = ws.cell(r, layout.total_col).value
        print(f"Row {r}: F={f!r} G={g!r} I={spec!r} K={k!r} M={m!r} N={n!r}")
        if str(f) != item["code"]:
            errors.append(f"row {r} product_no: {f!r} != {item['code']!r}")
        if g != item["quote_name"]:
            errors.append(f"row {r} quote_name mismatch")
        if k != item["qty"]:
            errors.append(f"row {r} qty: {k!r} != {item['qty']!r}")
        if m != item["unit_price"]:
            errors.append(f"row {r} unit_price: {m!r} != {item['unit_price']!r}")
        expected_total = round(item["unit_price"] * item["qty"], 2)
        if n != expected_total:
            errors.append(f"row {r} total: {n!r} != {expected_total!r}")

    # footer
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v and "Total Excluding" in str(v):
            footer_n = ws.cell(row, layout.totals_value_col).value
            expected_footer = round(sum(it["unit_price"] * it["qty"] for it in fill_items), 2)
            print(f"Footer row {row} N={footer_n!r} (expected {expected_footer})")
            if footer_n != expected_footer:
                errors.append(f"footer total: {footer_n!r} != {expected_footer!r}")
            break

    # quote date (K col)
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 9).value
        if v and "Tanggal Penawaran" in str(v):
            qdate = ws.cell(row, layout.quote_date_value_col or 11).value
            print(f"Quote date row {row} K={qdate!r}")
            if not qdate:
                errors.append("quote date not written")
            break

    wb.close()

    # Compare structure with filled reference (first data row cols)
    wb_ref = openpyxl.load_workbook(filled_ref, data_only=True)
    ws_ref = wb_ref.active
    ref_layout = _detect_quotation_layout(ws_ref)
    print("reference layout:", ref_layout.template_id)
    print("reference row 8 cols match layout:", ref_layout.product_no_col == VANTSING_LAYOUT.product_no_col)
    wb_ref.close()

    if errors:
        print("FAIL:")
        for e in errors:
            print(" -", e)
        return 1

    print("PASS: VANTSING fill test OK ->", work)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
