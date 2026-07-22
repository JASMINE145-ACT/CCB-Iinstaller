# -*- coding: utf-8 -*-
"""Structural validation for learn-by-data VANTSING smoke fixture."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from quotation.layout import VANTSING_LAYOUT
from quotation.quote_tools import TOTAL_ROW_MARKER, _cell_value

ROOT = Path(__file__).resolve().parents[2]
FIXTURE = ROOT / "data" / "smoke" / "learn-by-data-vantsing-filled.xlsx"
MANIFEST = ROOT / "data" / "smoke" / "learn-by-data-vantsing-filled.manifest.json"


def _find_total_row(ws) -> int | None:
    for row in ws.iter_rows():
        row_idx = row[0].row if row else 0
        for cell in row:
            if TOTAL_ROW_MARKER in _cell_value(cell):
                return row_idx
    return None


@pytest.fixture
def fixture_path() -> Path:
    if not FIXTURE.exists():
        pytest.skip(f"fixture missing: {FIXTURE} — run python/scripts/generate_learn_by_data_smoke_fixture.py")
    return FIXTURE


def test_learn_by_data_smoke_fixture_structure(fixture_path: Path) -> None:
    layout = VANTSING_LAYOUT
    wb = openpyxl.load_workbook(fixture_path, read_only=True, data_only=True)
    ws = wb.active
    total_row = _find_total_row(ws)
    assert total_row is not None
    assert total_row - layout.data_start_row >= 15

    data_rows = 0
    for row_idx in range(layout.data_start_row, total_row):
        name = str(ws.cell(row=row_idx, column=layout.inquiry_name_col).value or "").strip()
        spec = str(ws.cell(row=row_idx, column=layout.inquiry_spec_col).value or "").strip()
        code = str(ws.cell(row=row_idx, column=layout.product_no_col).value or "").strip()
        if name or spec:
            data_rows += 1
            assert code, f"row {row_idx} missing actual_code in col F"
    wb.close()
    assert data_rows >= 15


def test_learn_by_data_smoke_manifest_scenarios() -> None:
    if not MANIFEST.exists():
        pytest.skip(f"manifest missing: {MANIFEST} — run python/scripts/generate_learn_by_data_smoke_fixture.py")
    data = json.loads(MANIFEST.read_text(encoding="utf-8"))
    scenarios = data.get("scenarios") or {}
    assert scenarios.get("match", 0) >= 3
    assert scenarios.get("in-candidates", 0) >= 3
    assert scenarios.get("not-in-candidates", 0) >= 3
    assert scenarios.get("zero-candidate", 0) >= 1
    assert data.get("row_count", 0) >= 15


def test_deploy_skill_source_exists() -> None:
    # Skill moved to the vertical package (whitelist §8.2.1).
    skill = ROOT / "ccb-installer" / "packages" / "vertical" / "com.wanding.trade" / "skills" / "quotation-learn-by-data" / "SKILL.md"
    assert skill.is_file()
    text = skill.read_text(encoding="utf-8")
    assert "quotation-learn-by-data" in text
    assert "show_candidates" in text
    assert "VANTSING" in text
