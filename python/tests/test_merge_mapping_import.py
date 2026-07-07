"""Tests for merge_mapping_import script."""
from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path

import openpyxl

from quotation.learn_by_data_mapping import append_mapping_pending_row

ROOT = Path(__file__).resolve().parents[2]
MERGE_PATH = ROOT / "python" / "scripts" / "merge_mapping_import.py"


def _load_merge_module():
    spec = importlib.util.spec_from_file_location("merge_mapping_import", MERGE_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules["merge_mapping_import"] = module
    spec.loader.exec_module(module)
    return module


def _write_mapping_xlsx(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def test_merge_overwrites_same_keyword_when_allow_overwrite(tmp_path: Path, monkeypatch) -> None:
    merge_mapping_import = _load_merge_module()
    mapping_path = tmp_path / "mapping_table.xlsx"
    pending_path = tmp_path / "pending.jsonl"
    _write_mapping_xlsx(
        mapping_path,
        [
            ["询价货物名称", "询价规格型号", "产品编号", "报价名称"],
            ["50卷波纹管", "DN20", "8010062265", "old name"],
        ],
    )

    append_mapping_pending_row(
        {
            "inquiry_name": "50卷波纹管",
            "inquiry_spec": "DN20",
            "product_code": "8030020808",
            "quotation_name": "new name",
            "source_file": "q.xlsx",
            "source_sheet": "S1",
            "source_row": 10,
        },
        allow_overwrite=True,
        path=pending_path,
    )

    monkeypatch.setattr(merge_mapping_import, "invalidate_mapping_cache", lambda: None)

    result = merge_mapping_import.merge_pending(
        mapping_path=mapping_path,
        pending_path=pending_path,
        dry_run=False,
    )
    assert result["appended"] == 1
    assert result["overwritten"] == 1

    wb = openpyxl.load_workbook(mapping_path, read_only=True, data_only=True)
    ws = wb.active
    data_rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    assert len(data_rows) == 1
    assert str(data_rows[0][2]) == "8030020808"

    pending = [json.loads(line) for line in pending_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    assert pending[0]["status"] == "merged"
