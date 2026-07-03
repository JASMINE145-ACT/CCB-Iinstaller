# -*- coding: utf-8 -*-
"""Tests for VANTSING quotation Excel formulas (row total + footer SUM)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest

from quotation.layout import VANTSING_LAYOUT
from quotation.quote_tools import TOTAL_ROW_MARKER, _cell_value, fill_quotation

ROOT = Path(__file__).resolve().parents[2]
BLANK_TEMPLATE = ROOT / "data" / "空白标准报价单.xlsx"


def _find_total_row(ws) -> int | None:
    for row in ws.iter_rows():
        row_idx = row[0].row if row else 0
        for cell in row:
            if TOTAL_ROW_MARKER in _cell_value(cell):
                return row_idx
    return None


@pytest.fixture
def work_xlsx(tmp_path: Path) -> Path:
    if not BLANK_TEMPLATE.exists():
        pytest.skip(f"template missing: {BLANK_TEMPLATE}")
    out = tmp_path / "quotation_formulas.xlsx"
    shutil.copy2(BLANK_TEMPLATE, out)
    return out


def test_fill_10_rows_uses_row_and_footer_formulas(work_xlsx: Path) -> None:
    layout = VANTSING_LAYOUT
    fill_items = [
        {
            "row": 8 + i,
            "code": f"801000000{i}",
            "quote_name": f"Product {i + 1}",
            "specification": f"dn{50 + i}",
            "unit_price": 1000 + i * 50,
            "qty": 5,
            "inquiry_name": f"Inquiry {i + 1}",
        }
        for i in range(10)
    ]

    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))
    assert result["success"] is True

    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb.active
    total_row = _find_total_row(ws)
    assert total_row == 18

    assert ws.cell(row=8, column=layout.total_col).value == "=M8*K8"
    assert ws.cell(row=17, column=layout.total_col).value == "=M17*K17"
    assert ws.cell(row=total_row, column=layout.totals_value_col).value == "=SUM(N8:N17)"
    assert ws.cell(row=total_row + 1, column=layout.totals_value_col).value == "=N18*0.11"
    assert ws.cell(row=total_row + 3, column=layout.totals_value_col).value == "=SUM(N18:N20)"
    wb.close()


def test_wuhuo_row_total_is_zero_not_formula(work_xlsx: Path) -> None:
    layout = VANTSING_LAYOUT
    fill_items = [
        {
            "row": 8,
            "code": "无货",
            "quote_name": "Missing item",
            "unit_price": 1000,
            "qty": 2,
            "inquiry_name": "X",
        }
    ]

    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))
    assert result["success"] is True

    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb.active
    assert ws.cell(row=8, column=layout.total_col).value == 0
    wb.close()


def test_fill_return_value_still_python_sum(work_xlsx: Path) -> None:
    fill_items = [
        {
            "row": 8,
            "code": "8010000001",
            "quote_name": "P1",
            "unit_price": 100,
            "qty": 3,
            "inquiry_name": "I1",
        }
    ]
    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))
    assert result["success"] is True
    # fill_quotation does not return total in dict today - verify filled_count only
    assert result["filled_count"] == 1
