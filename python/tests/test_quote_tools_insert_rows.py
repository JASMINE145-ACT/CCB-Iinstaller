# -*- coding: utf-8 -*-
"""Tests for VANTSING quotation insert-rows beyond default 10 data lines."""
from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.cell.cell import MergedCell

from quotation.layout import VANTSING_LAYOUT
from quotation.quote_tools import (
    TOTAL_ROW_MARKER,
    _cell_value,
    fill_quotation,
)

ROOT = Path(__file__).resolve().parents[2]
BLANK_TEMPLATE = ROOT / "data" / "空白标准报价单.xlsx"


def _find_total_row(ws) -> int | None:
    for row in ws.iter_rows():
        row_idx = row[0].row if row else 0
        for cell in row:
            if TOTAL_ROW_MARKER in _cell_value(cell):
                return row_idx
    return None


def _make_fill_items(count: int, *, start_row: int = 8) -> list[dict]:
    items: list[dict] = []
    for i in range(count):
        items.append(
            {
                "row": start_row + i,
                "code": f"801000000{i}",
                "quote_name": f"Test Product {i + 1}",
                "specification": f"dn{25 + i * 5}",
                "unit_price": 1000 + i * 100,
                "qty": 10,
                "inquiry_name": f"Inquiry {i + 1}",
                "satuan": "pcs",
                "brand": "LESSO",
            }
        )
    return items


@pytest.fixture
def work_xlsx(tmp_path: Path) -> Path:
    if not BLANK_TEMPLATE.exists():
        pytest.skip(f"template missing: {BLANK_TEMPLATE}")
    out = tmp_path / "quotation_insert_rows.xlsx"
    shutil.copy2(BLANK_TEMPLATE, out)
    return out


def test_fill_12_rows_writes_quote_columns_on_inserted_rows(work_xlsx: Path) -> None:
    layout = VANTSING_LAYOUT
    fill_items = _make_fill_items(12)
    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))
    assert result["success"] is True
    assert result["filled_count"] == 12

    out_path = Path(result["output_path"])
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    total_row = _find_total_row(ws)
    assert total_row == 20

    for row in (18, 19):
        for col in (layout.product_no_col, layout.quote_name_col, layout.quote_spec_col, layout.total_col):
            cell = ws.cell(row=row, column=col)
            assert not isinstance(cell, MergedCell), f"row {row} col {col} is MergedCell"
            assert cell.value is not None, f"row {row} col {col} is empty"

        code = ws.cell(row=row, column=layout.product_no_col).value
        assert str(code).startswith("801000000")
        row_formula = ws.cell(row=row, column=layout.total_col).value
        assert row_formula == f"=M{row}*K{row}"

    footer_total = ws.cell(row=total_row, column=layout.totals_value_col).value
    assert footer_total == "=SUM(N8:N19)"
    assert ws.cell(row=total_row + 1, column=layout.totals_value_col).value == f"=N{total_row}*0.11"
    assert ws.cell(row=total_row + 3, column=layout.totals_value_col).value == f"=SUM(N{total_row}:N{total_row + 2})"
    wb.close()


def test_fill_10_rows_does_not_shift_total_row(work_xlsx: Path) -> None:
    fill_items = _make_fill_items(10)
    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))
    assert result["success"] is True

    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb.active
    assert _find_total_row(ws) == 18
    wb.close()


def test_vantsing_fill_writes_remark_to_catatan_column(work_xlsx: Path) -> None:
    layout = VANTSING_LAYOUT
    fill_items = _make_fill_items(1)
    fill_items[0]["remark"] = "HENG XIN INTERNATIONAL INDONESIA"

    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))
    assert result["success"] is True

    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb.active
    assert ws.cell(row=8, column=layout.remark_col).value == "HENG XIN INTERNATIONAL INDONESIA"
    wb.close()


def test_fill_11_rows_inserts_one_row(work_xlsx: Path) -> None:
    layout = VANTSING_LAYOUT
    fill_items = _make_fill_items(11)
    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))
    assert result["success"] is True

    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb.active
    total_row = _find_total_row(ws)
    assert total_row == 19

    row = 18
    assert not isinstance(ws.cell(row, layout.product_no_col), MergedCell)
    assert ws.cell(row, layout.product_no_col).value is not None
    wb.close()


def test_no_stray_merge_on_inserted_data_rows(work_xlsx: Path) -> None:
    fill_items = _make_fill_items(12)
    result = fill_quotation(str(work_xlsx), fill_items, output_path=str(work_xlsx))

    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb.active
    total_row = _find_total_row(ws)
    assert total_row is not None
    last_data_row = total_row - 1
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= 8 and merged_range.max_row <= last_data_row:
            pytest.fail(f"unexpected merge on data rows: {merged_range.coord}")
    wb.close()
