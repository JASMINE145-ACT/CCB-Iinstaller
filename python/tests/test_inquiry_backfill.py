"""Inquiry column (B–E) backfill when empty — agent-provided fields."""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

PYTHON_ROOT = Path(__file__).resolve().parent.parent
if str(PYTHON_ROOT) not in sys.path:
    sys.path.insert(0, str(PYTHON_ROOT))

import openpyxl  # noqa: E402

from quotation.inquiry_backfill import (  # noqa: E402
    backfill_inquiry_columns_if_empty,
    resolve_inquiry_fields_from_item,
)
from quotation.layout import VANTSING_LAYOUT  # noqa: E402
from quotation.quote_tools import (  # noqa: E402
    _detect_quotation_layout,
    fill_quotation,
)


class TestResolveInquiryFields(unittest.TestCase):
    def test_prefers_explicit_inquiry_fields(self) -> None:
        item = {
            "inquiry_name": "直接50",
            "inquiry_spec": "dn50",
            "source_keyword": "别的",
            "specification": "dn110",
        }
        fields = resolve_inquiry_fields_from_item(item)
        self.assertEqual(fields["inquiry_name"], "直接50")
        self.assertEqual(fields["inquiry_spec"], "dn50")

    def test_falls_back_to_source_keyword_and_specification(self) -> None:
        item = {
            "source_keyword": "直接50",
            "specification": "dn50",
            "quote_name": "PVC-U 直接 dn50",
        }
        fields = resolve_inquiry_fields_from_item(item)
        self.assertEqual(fields["inquiry_name"], "直接50")
        self.assertEqual(fields["inquiry_spec"], "dn50")


class TestBackfillInquiryColumns(unittest.TestCase):
    def test_fills_empty_bcde_only(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        layout = VANTSING_LAYOUT
        row = 8
        ws.cell(row, layout.inquiry_name_col, value="已有名称")
        item = {
            "inquiry_name": "直接50",
            "inquiry_spec": "dn50",
            "satuan": "个",
        }

        backfill_inquiry_columns_if_empty(ws, row, layout, item, satuan="个", qty=10)

        self.assertEqual(ws.cell(row, layout.inquiry_name_col).value, "已有名称")
        self.assertEqual(ws.cell(row, layout.inquiry_spec_col).value, "dn50")
        self.assertEqual(ws.cell(row, layout.inquiry_unit_col).value, "个")
        self.assertEqual(ws.cell(row, layout.inquiry_qty_col).value, 10)

    def test_skips_when_all_filled(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        layout = VANTSING_LAYOUT
        row = 8
        ws.cell(row, layout.inquiry_name_col, value="A")
        ws.cell(row, layout.inquiry_spec_col, value="B")
        ws.cell(row, layout.inquiry_unit_col, value="C")
        ws.cell(row, layout.inquiry_qty_col, value=1)

        backfill_inquiry_columns_if_empty(
            ws,
            row,
            layout,
            {"inquiry_name": "新", "inquiry_spec": "dn50"},
            satuan="个",
            qty=99,
        )

        self.assertEqual(ws.cell(row, layout.inquiry_name_col).value, "A")
        self.assertEqual(ws.cell(row, layout.inquiry_spec_col).value, "B")
        self.assertEqual(ws.cell(row, layout.inquiry_unit_col).value, "C")
        self.assertEqual(ws.cell(row, layout.inquiry_qty_col).value, 1)

    def test_specification_fallback_for_inquiry_spec_column(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        layout = VANTSING_LAYOUT
        row = 8
        item = {
            "source_keyword": "直接50",
            "specification": "dn50",
            "quote_name": "PVC-U 直接 dn50",
        }

        backfill_inquiry_columns_if_empty(ws, row, layout, item, satuan=None, qty=None)

        self.assertEqual(ws.cell(row, layout.inquiry_name_col).value, "直接50")
        self.assertEqual(ws.cell(row, layout.inquiry_spec_col).value, "dn50")


class TestFillQuotationInquiryBackfill(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        data_dir = PYTHON_ROOT.parent / "data"
        matches = [p for p in data_dir.glob("*.xlsx") if "空白" in p.name]
        if not matches:
            raise unittest.SkipTest("blank VANTSING template xlsx not found under data/")
        cls.blank_template = matches[0]

    def test_fill_backfills_inquiry_name_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp) / "quote.xlsx"
            import shutil

            shutil.copy(self.blank_template, work)

            row = 8
            result = fill_quotation(
                str(work),
                [
                    {
                        "row": row,
                        "code": "8010071492",
                        "quote_name": "PVC-U 直接 dn50",
                        "unit_price": 8500,
                        "qty": 1,
                        "specification": "dn50",
                        "inquiry_name": "直接50",
                    }
                ],
                output_path=str(Path(tmp) / "filled.xlsx"),
            )
            self.assertTrue(result.get("success"), result.get("error"))

            wb = openpyxl.load_workbook(result["output_path"], data_only=True)
            ws = wb.active
            layout = _detect_quotation_layout(ws)

            self.assertEqual(ws.cell(row, layout.inquiry_name_col).value, "直接50")
            self.assertEqual(ws.cell(row, layout.inquiry_spec_col).value, "dn50")

    def test_fill_backfills_from_agent_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp) / "quote.xlsx"
            import shutil

            shutil.copy(self.blank_template, work)

            row = 8
            result = fill_quotation(
                str(work),
                [
                    {
                        "row": row,
                        "code": "8010071492",
                        "quote_name": "PVC-U 直接 dn50",
                        "unit_price": 8500,
                        "qty": 10,
                        "specification": "dn50",
                        "inquiry_name": "直接50",
                        "inquiry_spec": "dn50",
                        "satuan": "个",
                    }
                ],
                output_path=str(Path(tmp) / "filled.xlsx"),
            )
            self.assertTrue(result.get("success"), result.get("error"))

            wb = openpyxl.load_workbook(result["output_path"], data_only=True)
            ws = wb.active
            layout = _detect_quotation_layout(ws)

            self.assertEqual(ws.cell(row, layout.inquiry_name_col).value, "直接50")
            self.assertEqual(ws.cell(row, layout.inquiry_spec_col).value, "dn50")
            self.assertEqual(ws.cell(row, layout.inquiry_unit_col).value, "个")
            self.assertEqual(ws.cell(row, layout.inquiry_qty_col).value, 10)


if __name__ == "__main__":
    unittest.main()
