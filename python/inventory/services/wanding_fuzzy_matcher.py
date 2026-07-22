"""
万鼎价格库匹配 - DataBase-style 模糊逻辑

仅此一种查询逻辑：token + 同义词扩展 + 规格等价 + score 排序。
借鉴 DataBase- 项目 search_with_keywords(strict=False, return_score=True)。
"""

from __future__ import annotations

import logging
import re
import threading
from pathlib import Path
from typing import Any, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# 价格列（0-based）：与价格库表头顺序一致；多档位共用同一列时取同一索引
PRICE_COLS = {
    "FACTORY_INC_TAX": 4,
    "FACTORY_EXC_TAX": 5,
    "PURCHASE_EXC_TAX": 6,
    # 下列索引基于「万鼎价格库_管材与国标管件_标准格式.xlsx」表头（含 Product_Type 列后偏移 +1）：
    # 列5-7: 出厂价；列8/9: A 档 利润率/报单价格；列10/11: B 档；列12/13: C 档；列14/15: D 档；列16/17: D 低利润率；列18/19: E 档。
    "A_MARGIN": 9,
    "A_QUOTE": 9,
    "B_MARGIN": 11,
    "B_QUOTE": 11,
    "C_MARGIN": 13,
    "C_QUOTE": 13,
    "D_MARGIN": 15,
    "D_QUOTE": 15,
    "D_LOW": 17,
    "E_MARGIN": 19,
    "E_QUOTE": 19,
    # 兼容旧代码
    "A": 9,
    "A_TURN": 9,
    "A_ANNUAL": 9,
    "B": 11,
    "B_TURN": 11,
    "B_ANNUAL": 11,
    "B_QUOTE": 11,
    "C": 13,
    "C_TURN": 13,
    "C_QUOTE": 13,
    "D": 15,
    "D_NOADJ": 15,
    "D_WHOLESALE": 17,
    "E": 19,
}

# 对应每个档位价格列的「利润率」列索引（0-based）。
PROFIT_COLS: dict[str, int] = {
    "A_MARGIN": 8,
    "A_QUOTE": 8,
    "B_MARGIN": 10,
    "B_QUOTE": 10,
    "C_MARGIN": 12,
    "C_QUOTE": 12,
    "D_MARGIN": 14,
    "D_QUOTE": 14,
    "D_LOW": 16,
    "E_MARGIN": 18,
    "E_QUOTE": 18,
}


def _normalize_price_level(customer_level: str) -> str:
    """将用户/Agent 传入的档位或价格类型统一为 PRICE_COLS 的 key。"""
    s = (customer_level or "B_QUOTE").strip()
    if not s:
        return "B_QUOTE"
    compact = re.sub(r"[\s_\-（）()：:]+", "", s).lower()
    # 中文：出厂价含税/不含税、采购不含税
    if "出厂价" in s and "不含税" in s:
        return "FACTORY_EXC_TAX"
    if "出厂价" in s and "含税" in s:
        return "FACTORY_INC_TAX"
    if "采购" in s and "不含税" in s:
        return "PURCHASE_EXC_TAX"
    # LOCAL / RUCIKA / PE 须在通用 exctax/inctax 之前，避免 localexctax 等被误判为出厂价
    if any(alias in compact for alias in ("localexctax", "local不含税", "localexc")) or (
        "local" in compact and ("不含税" in s or "exctax" in compact)
    ):
        return "LOCAL_EXC_TAX"
    if any(alias in compact for alias in ("localinctax", "local含税", "localinc")) or (
        "local" in compact and ("inctax" in compact or ("含税" in s and "不含税" not in s))
    ):
        return "LOCAL_INC_TAX"
    if "local" in compact or "本地价" in s or "本地价格" in s:
        return "LOCAL_EXC_TAX"
    if "rucika" in compact or "rucika" in s.lower():
        if any(alias in compact for alias in ("quote2", "报单2", "第二组")):
            return "RUCIKA_QUOTE_2"
        if any(alias in compact for alias in ("quote1", "报单1", "第一组")):
            return "RUCIKA_QUOTE_1"
        if any(alias in compact for alias in ("pricelistexc", "excvat11")) or (
            "pricelist" in compact and compact.endswith("exc")
        ) or ("目录价" in s and "不含税" in s):
            return "RUCIKA_PRICELIST_EXC"
        if any(alias in compact for alias in ("pricelistinc", "incvat11")) or (
            "pricelist" in compact and compact.endswith("inc")
        ) or ("目录价" in s and "含税" in s and "不含税" not in s):
            return "RUCIKA_PRICELIST_INC"
        return "RUCIKA_QUOTE_1"
    if any(alias in compact for alias in ("penominal", "pe面价", "nominalprice")) or (
        "pe" in compact and "面价" in s
    ):
        return "PE_NOMINAL"
    if any(alias in compact for alias in ("pefactory", "pe出厂价", "pe出厂")):
        return "PE_FACTORY"
    if any(alias in compact for alias in ("includetax", "含税出厂价", "出厂价含税")):
        return "FACTORY_INC_TAX"
    if any(alias in compact for alias in ("excludetax", "exctax", "不含税出厂价", "出厂价不含税")):
        return "FACTORY_EXC_TAX"
    if any(alias in compact for alias in ("purchaseprice", "purchaseexctax", "采购价", "采购价格", "采购不含税")):
        return "PURCHASE_EXC_TAX"
    if any(alias in compact for alias in ("二级代理", "二级代理a", "a级别", "a级", "aprice", "pricea")):
        return "A_QUOTE"
    if any(alias in compact for alias in ("一级代理", "一级代理b", "b级别", "b级", "默认价", "默认价格", "generalprice", "priceb")):
        return "B_QUOTE"
    if any(alias in compact for alias in ("聚万大客户", "聚万", "c级别", "c级", "pricec")):
        return "C_QUOTE"
    if any(alias in compact for alias in ("青山大客户降低", "青山降低", "d低", "dlow", "price_d_low", "pricedlow")):
        return "D_LOW"
    if any(alias in compact for alias in ("青山大客户", "青山价格", "青山价", "青山", "d级别", "d级", "mrkongprice", "孔总价", "孔总价格", "孔总", "priced")):
        return "D_QUOTE"
    if any(alias in compact for alias in ("大唐大客户", "大唐价格", "大唐价", "大唐", "e级别", "e级", "pricee")):
        return "E_QUOTE"
    if any(alias in compact for alias in ("其他客户价", "其他价格", "其他价", "其他")):
        return "B_QUOTE"
    # 英文/代码
    u = s.upper().replace(" ", "_")
    if u in ("D_LOW", "D LOW", "DLOW"):
        return "D_LOW"
    for key in (
        "FACTORY_INC_TAX",
        "FACTORY_EXC_TAX",
        "PURCHASE_EXC_TAX",
        "LOCAL_EXC_TAX",
        "LOCAL_INC_TAX",
        "RUCIKA_PRICELIST_EXC",
        "RUCIKA_PRICELIST_INC",
        "RUCIKA_QUOTE_1",
        "RUCIKA_QUOTE_2",
        "PE_NOMINAL",
        "PE_FACTORY",
    ):
        if u == key or u.replace("_", "") == key.replace("_", ""):
            return key
    single_letter_quote = {"A": "A_QUOTE", "B": "B_QUOTE", "C": "C_QUOTE", "D": "D_QUOTE", "E": "E_QUOTE"}
    if u in single_letter_quote:
        return single_letter_quote[u]
    return u if u in PRICE_COLS else "B_QUOTE"


# 档位代码 → 全名（与价格库表头一致，Chat/Work 统一显示）
PRICE_LEVEL_DISPLAY_NAMES: dict[str, str] = {
    "FACTORY_INC_TAX": "出厂价_含税",
    "FACTORY_EXC_TAX": "出厂价_不含税",
    "PURCHASE_EXC_TAX": "采购不含税",
    "A_MARGIN": "（二级代理）A级别 利润率",
    "A_QUOTE": "（二级代理）A级别 报单价格",
    "B_MARGIN": "（一级代理）B级别 利润率",
    "B_QUOTE": "（一级代理）B级别 报单价格",
    "C_MARGIN": "（聚万大客户）C级别 利润率",
    "C_QUOTE": "（聚万大客户）C级别报单价格",
    "D_MARGIN": "（青山大客户）D级别 利润率",
    "D_QUOTE": "（青山大客户）D级别 报单价格",
    "D_LOW": "（青山大客户）D级别 降低利润率",
    "E_MARGIN": "（大唐大客户）E级别（包运费） 利润率",
    "E_QUOTE": "（大唐大客户）E级别（包运费） 报单价格",
    "LOCAL_EXC_TAX": "LOCAL EXC TAX（本地不含税价）",
    "LOCAL_INC_TAX": "LOCAL INC TAX（本地含税价）",
    "RUCIKA_PRICELIST_EXC": "RUCIKA Pricelist Exc Vat 11%",
    "RUCIKA_PRICELIST_INC": "RUCIKA Pricelist Inc Vat 11%",
    "RUCIKA_QUOTE_1": "RUCIKA 报单价格（第一组）",
    "RUCIKA_QUOTE_2": "RUCIKA 报单价格（第二组）",
    "PE_NOMINAL": "PE 面价（印尼盾/条）",
    "PE_FACTORY": "PE 出厂价/条",
    # 兼容旧代码
    "A": "（二级代理）A级别 利润率",
    "A_TURN": "（二级代理）A级别 利润率",
    "A_ANNUAL": "（二级代理）A级别 报单价格",
    "B": "（一级代理）B级别 利润率",
    "B_TURN": "（一级代理）B级别 利润率",
    "B_ANNUAL": "（一级代理）B级别 报单价格",
    "C": "（聚万大客户）C级别 利润率",
    "C_TURN": "（聚万大客户）C级别 利润率",
    "D": "（青山大客户）D级别 利润率",
    "D_NOADJ": "（青山大客户）D级别 报单价格",
    "D_WHOLESALE": "（青山大客户）D级别 降低利润率",
    "E": "（大唐大客户）E级别（包运费） 利润率",
}


def get_price_level_display_name(customer_level: str) -> str:
    """将档位代码转为全名，供界面与接口返回使用。"""
    key = _normalize_price_level(customer_level or "B_QUOTE")
    return PRICE_LEVEL_DISPLAY_NAMES.get(key, key)


SYNONYM_GROUPS = [
    {"直接", "直接头", "直通", "直通接头"},
    {"变径", "异径"},
    {"大小头", "异径直通", "异径套", "变径直接", "异径直接"},
    {"内丝", "内螺纹"}, {"外丝", "外螺纹"},
    {"锁母", "锁扣", "管接头"},
    {"止回阀", "截止阀"},
    {"穿线管", "电线管"},
    {"半弯", "弯头"},
    {"承插", "承插式"},
    {"堵头", "管帽"},
]

# 模块级预计算，避免每次调用时重建（SYNONYM_GROUPS 不变时永远有效）
_SYNONYM_TO_GROUP: dict[str, frozenset] = {
    syn: frozenset(group) for group in SYNONYM_GROUPS for syn in group
}
_SORTED_SYNONYMS: list[str] = sorted(_SYNONYM_TO_GROUP.keys(), key=len, reverse=True)

# 单字 token（如「三」「通」）在打分中的权重，相对于多字 token 的 1.0
_SINGLE_CHAR_WEIGHT = 0.5

# 询价关键词中的英文/印尼语/口语 → 中文品名，用于筛选时命中库内品名（与 wanding_business_knowledge.md 保持一致，便于 LLM 选型有思路）
QUERY_TERM_TO_CHINESE = [
    ("4 cabang", "管四通圆接线盒"),
    ("conduit", "电线管"), ("counduit", "电线管"), ("pipa", "管"),
    ("socket", "管直通"), ("套筒", "管直通"),
    ("klem", "管夹"),
    ("cabang", "四通"), ("tdust", "四通"),
    ("热熔器", "焊接机"), ("热熔机", "焊接机"), ("熔接器", "焊接机"),
    ("四通接线盒", "管四通圆接线盒"),
    ("马鞍卡", "管夹"),
    # AW 在印尼管材场景下常指 AW 给水系列
    ("aw", "给水 aw给水系列"),
]


_UNICODE_FRACTIONS = {
    "½": "1/2",
    "¾": "3/4",
    "¼": "1/4",
}


def _normalize_unicode_fractions(keywords: str) -> str:
    s = keywords or ""
    for uchar, ascii_frac in _UNICODE_FRACTIONS.items():
        s = s.replace(uchar, ascii_frac)
    return s


def _is_ceiling_drat(norm_text: str) -> bool:
    return _contains_any(norm_text, [r"\bsteel\s*drat\b", r"\bstelldrat\b"])


def _is_fitting_drat_thread(norm_text: str) -> bool:
    """Elbow drat / 丝扣弯头 in AW/fitting context — not ceiling Stelldrat."""
    if _is_ceiling_drat(norm_text):
        return False
    if not _contains_any(norm_text, [r"\bdrat\b", r"丝扣", r"螺纹弯头", r"螺纹弯"]):
        return False
    if _query_fitting(norm_text) in {"elbow", "tee", "coupling", "reducer", "fitting"}:
        return True
    return _contains_any(norm_text, [r"\baw\b", r"给水"])


def _apply_drat_thread_expansion(keywords: str) -> str:
    norm = _normalize(keywords or "")
    if not _is_fitting_drat_thread(norm):
        return keywords
    extra = " 内螺纹 丝扣弯头 螺纹弯头"
    if extra.strip() not in norm:
        return f"{keywords} {extra}".strip()
    return keywords


def _normalize_keyword_terms(keywords: str) -> str:
    """将询价中的英文/印尼语替换为中文品名词，便于筛选命中库内品名。"""
    s = (keywords or "").strip()
    for eng, ch in QUERY_TERM_TO_CHINESE:
        s = re.sub(r"\b" + re.escape(eng) + r"\b", ch, s, flags=re.I)
    return s.strip()


# 询价意图词（价格/档位）不应参与产品字段匹配，否则会把正确候选过滤掉
_QUERY_INTENT_STOPWORDS = {
    "报价", "报单", "价格", "价", "报价价格",
    "一级", "二级", "三级", "代理", "代理价", "一级代理", "二级代理",
    "a级", "b级", "c级", "d级", "e级",
}


def _strip_query_intent_terms(keywords: str) -> str:
    """移除询价中的非品名意图词，保留材质/规格/品类 token 用于字段匹配。"""
    s = _normalize(keywords or "")
    if not s:
        return ""
    # 先按长词优先剔除，避免残留碎片（如先去「一级代理」再去「一级」）
    for term in sorted(_QUERY_INTENT_STOPWORDS, key=len, reverse=True):
        s = re.sub(re.escape(term), " ", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# 业务知识中【字段匹配同义与规格】规则缓存，供字段匹配阶段使用（与 LLM 选型共用同一 knowledge 文件）
_FIELD_MATCHING_RULES_CACHE: dict = {}  # {"path": str, "mtime": float|None, "rules": [(sources, targets), ...]}


def _is_field_matching_section_header(line: str) -> bool:
    """True for 【字段匹配…】 or markdown H2 containing 字段匹配同义."""
    if "【字段匹配" in line:
        return True
    if "字段匹配同义" in line and line.lstrip().startswith("#"):
        return True
    return False


def _parse_field_matching_sources(left: str) -> List[str]:
    """
    Parse the left-hand source list of a field-matching rule.

    - No `/`: space-separated tokens are OR alternatives (`elbow 弯` → elbow|弯).
    - With `/`: each `/`-separated segment is one phrase (may contain spaces).
      Example: `elbow drat / drat` → ["elbow drat", "drat"] — not bare "elbow".
      (WANd.MATCH.FIELD_RULE_PARSE.001)
    """
    left = (left or "").strip()
    if not left:
        return []
    if "/" in left:
        return [p.strip() for p in left.split("/") if p.strip()]
    return [t.strip() for t in left.split() if t.strip()]


def _parse_field_matching_rules_from_content(content: str) -> List[tuple[List[str], List[str]]]:
    """
    Parse synonym expansion rules from knowledge markdown body.
    Section start: line with 【字段匹配…】 OR markdown heading with 字段匹配同义.
    Section end: next 【…】 heading OR next markdown H2 that is not the synonym header.
    """
    rules: List[tuple[List[str], List[str]]] = []
    in_section = False
    for raw in content.splitlines():
        line = raw.strip()
        if _is_field_matching_section_header(line):
            in_section = True
            continue
        if in_section:
            if line.startswith("【") and "字段匹配" not in line:
                break
            if line.startswith("## ") and not _is_field_matching_section_header(line):
                break
        if not in_section:
            continue
        # 解析 "- 源词 源词 → 检索词 检索词" 或 "  - ... → ..."
        if line.startswith("-"):
            line = line.lstrip("-").strip()
        if "→" in line:
            left, _, right = line.partition("→")
        elif "->" in line:
            left, _, right = line.partition("->")
        else:
            continue
        sources = _parse_field_matching_sources(left)
        targets = [t.strip() for t in right.split() if t.strip()]
        if sources and targets:
            rules.append((sources, targets))
    return rules


def _load_field_matching_rules_from_knowledge() -> List[tuple[List[str], List[str]]]:
    """
    从 wanding_business_knowledge.md 的【字段匹配同义与规格】段落解析规则，
    用于字段匹配阶段同义扩展，提高命中率。返回 [(source_terms, target_terms), ...]。
    """
    global _FIELD_MATCHING_RULES_CACHE
    try:
        from inventory.config import config
        from admin.org_knowledge_client import load_doc_content

        path_str = getattr(config, "WANDING_BUSINESS_KNOWLEDGE_PATH", None)
        content = load_doc_content(
            "wanding_business_knowledge",
            fallback_path=path_str,
            use_cache=True,
        )
        if not content and path_str:
            p = Path(path_str)
            if p.exists():
                content = p.read_text(encoding="utf-8")
        if not content:
            return []

        cache_key = "org-api:wanding_business_knowledge" if content else str(path_str)
        if _FIELD_MATCHING_RULES_CACHE.get("path") == cache_key:
            return _FIELD_MATCHING_RULES_CACHE.get("rules") or []

        rules = _parse_field_matching_rules_from_content(content)
        _FIELD_MATCHING_RULES_CACHE = {"path": cache_key, "mtime": None, "rules": rules}
        return rules
    except Exception as e:
        logger.debug("加载字段匹配规则失败: %s", e)
        return []


def _apply_knowledge_expansion(keywords: str) -> str:
    """
    根据业务知识【字段匹配同义与规格】规则，在字段匹配前扩展询价词，
    使口语/同义词能命中库内品名（如 直接→直通 排水、热熔器→焊接机）。
    """
    if not (keywords or "").strip():
        return keywords
    rules = _load_field_matching_rules_from_knowledge()
    added: List[str] = []
    kw_lower = (keywords or "").lower()
    for sources, targets in rules:
        for src in sources:
            if src.lower() in kw_lower or re.search(re.escape(src), keywords, re.I):
                added.extend(targets)
                break
    if not added:
        return keywords.strip()
    return (keywords.strip() + " " + " ".join(added)).strip()


# PN ↔ MPa 双向扩展
# PN 是公称压力等级：PN10 = 1.0MPa, PN16 = 1.6MPa, PN12.5 = 1.25MPa（数字 × 0.1 ≈ MPa）
# PN 前需要词边界\b；MPa 前数字后不需要严格\b（因为数字前可能是空格/汉字/标点）
_PN_RE = re.compile(r'\bPN\s*(\d+(?:\.\d+)?)(?![\da-zA-Z_])', re.IGNORECASE)
# 数字+MPa，尾部用负前瞻排除纯 ASCII 字母数字后缀（如 MPaA、MPa1），汉字/空格/结尾均放行
_MPA_RE = re.compile(r'(\d+(?:\.\d+)?)\s*MPA(?![\da-zA-Z])', re.IGNORECASE)


def _format_pressure_value(value: float) -> str:
    """格式化压力值，去除浮点尾巴（如 1.25 而非 1.2500000001）"""
    formatted = f"{value:.2f}".rstrip('0').rstrip('.')
    return formatted


def _apply_pressure_expansion(keywords: str) -> str:
    """
    双向扩展 PN ↔ MPa，并做数值格式化。
    - PN -> MPa：PN16 -> 1.6MPa
    - MPa -> PN：1.25MPa -> PN12.5
    - 格式化：1.60MPa / 1.6 MPa / 1.6mpa -> 统一 1.6MPa
    - 去重：已扩展过的等价形式不重复追加
    """
    if not (keywords or "").strip():
        return ""

    # 预扫描已存在的 PN/MPa 数值，避免重复追加
    seen_pn: set[str] = set()
    seen_mpa: set[str] = set()

    def _scan_pn(m: re.Match) -> str:
        pn_val = float(m.group(1))
        seen_pn.add(_format_pressure_value(pn_val))
        return m.group(0)

    def _scan_mpa(m: re.Match) -> str:
        mpa_val = float(m.group(1))
        seen_mpa.add(_format_pressure_value(mpa_val))
        return m.group(0)

    _PN_RE.sub(_scan_pn, keywords)
    _MPA_RE.sub(_scan_mpa, keywords)

    additions: list[str] = []

    # PN -> MPa：仅当原词中不存在等价 MPa 值
    def _sub_pn_to_mpa(m: re.Match) -> str:
        pn_val = float(m.group(1))
        mpa_val = pn_val * 0.1
        mpa_formatted = _format_pressure_value(mpa_val)
        if mpa_formatted in seen_mpa:
            return m.group(0)
        seen_mpa.add(mpa_formatted)
        additions.append(f"{mpa_formatted}MPa")
        return m.group(0)

    # MPa -> PN：仅当原词中不存在等价 PN 值
    def _sub_mpa_to_pn(m: re.Match) -> str:
        mpa_val = float(m.group(1))
        pn_val = mpa_val * 10
        pn_formatted = _format_pressure_value(pn_val)
        if pn_formatted in seen_pn:
            return m.group(0)
        seen_pn.add(pn_formatted)
        additions.append(f"PN{pn_formatted}")
        return m.group(0)

    result = keywords
    result = _PN_RE.sub(_sub_pn_to_mpa, result)
    result = _MPA_RE.sub(_sub_mpa_to_pn, result)

    if additions:
        result = result + " " + " ".join(additions)

    return result


MM_TO_INCH = {
    "16": '1/2"', "20": '3/4"', "25": '1"', "32": '1-1/4"', "40": '1-1/2"',
    "50": '2"', "65": '2-1/2"', "75": '3"', "100": '4"', "125": '5"',
    "150": '6"', "200": '8"', "250": '10"', "300": '12"',
}
INCH_TO_MM = {v: k for k, v in MM_TO_INCH.items()}

# 国内口语「N 分」管径 → 公称 DN 数字（与价格库 Describrition 中 DN15/DN20 对齐）
# 4 分 ≈ 1/2" → DN15；6 分 ≈ 3/4" → DN20（避免把「4」单拆出来导致与 DN15 无法匹配）
FEN_TO_MM_STR = {
    "4": "15",
    "6": "20",
}
# 整数「N 寸」→ DN（常用对照，与 _split_tokens 中 \d+寸 整段提取配合）
CUN_INTEGER_TO_MM_STR = {
    "1": "25",
    "2": "50",
    "3": "80",
    "4": "100",
    "5": "125",
    "6": "150",
    "8": "200",
    "10": "250",
    "12": "300",
}

# 脚本/旧测试兼容别名（仅「寸」映射；「分」见 FEN_TO_MM_STR）
CUN_TO_MM = CUN_INTEGER_TO_MM_STR

# 主径×副径口语规则：内外丝类询价里副径常按英寸侧理解。
THREAD_SIDE_TO_INCH = {
    "15": '1/2"',
    "16": '1/2"',
    "20": '1/2"',
    "25": '3/4"',
    "32": '1"',
    "40": '1-1/4"',
    "50": '1-1/2"',
}

# 日标 PVC-U 排水管件：外径口语数字 → 公称通径 DN
# 这些数字不在 MM_TO_INCH（标准 DN 系列）中，出现在查询里必然是外径，不是 DN：
#   OD63  → DN50 (2")   OD110 → DN100 (4")   OD160 → DN150 (6")
# 添加 DN 等价后，日标产品（Describrition 含 DN100/DN150）才能进入候选；
# 原 OD 值同时保留，让国标产品（Describrition 含 dn110/dn160）也能命中，
# 最终由 LLM 或 Product_Type 过滤在两类候选中选出正确结果。
OD_TO_DN_JIS: dict[str, str] = {
    "63": "50",    # OD63  → DN50  (2")
    "110": "100",  # OD110 → DN100 (4")
    "160": "150",  # OD160 → DN150 (6")
}


def _normalize(s: str) -> str:
    s = (s or "").lower().strip()
    s = s.replace("－", "-").replace("—", "-").replace("（", "(").replace("）", ")")
    s = re.sub(r"[_\t]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _normalize_chinese_number_order(s: str) -> str:
    """
    Keep legacy token normalization behavior for tests and callers:
    reorder "50三通" -> "三通50" while preserving more specific phrases.
    """
    text = s or ""
    m = re.fullmatch(r"(\d+)([\u4e00-\u9fff]{2})", text)
    if m:
        return f"{m.group(2)}{m.group(1)}"
    return text


def _get_synonym_words(word: str) -> frozenset:
    return _SYNONYM_TO_GROUP.get(word, frozenset({word}))


def _expand_unit_tokens(token: str, material: Optional[str] = None) -> set:
    eqs = {token}
    # 支持分数英寸规格（如 3/4、1-1/4），统一补全为带引号形式参与映射
    if re.fullmatch(r"\d+(?:-\d+)?/\d+", token):
        token = token + '"'
        eqs.add(token)
    # 行业场景下将 DE 视为 DN 的等价写法（如 DE50≈DN50）
    if token.startswith("de"):
        num = token[2:]
        if not num.isdigit():
            return eqs
        if num in MM_TO_INCH:
            eqs.add(MM_TO_INCH[num])
        eqs.add(num)
        eqs.add("dn" + num)
        return eqs
    if token.startswith("dn"):
        num = token[2:]
        if num in MM_TO_INCH:
            eqs.add(MM_TO_INCH[num])
        eqs.add(num)
        return eqs
    if token.isdigit() and token in MM_TO_INCH:
        eqs.add("dn" + token)
        eqs.add(MM_TO_INCH[token])
        return eqs
    # 日标外径 → DN 等价（63→50, 110→100, 160→150）
    # 原 OD 值（token）已在 eqs 中，保留以命中国标产品；DN 等价用于命中日标产品
    if token.isdigit() and token in OD_TO_DN_JIS:
        dn_num = OD_TO_DN_JIS[token]
        eqs.add(dn_num)
        eqs.add("dn" + dn_num)
        if dn_num in MM_TO_INCH:
            eqs.add(MM_TO_INCH[dn_num])
        return eqs
    if token in INCH_TO_MM:
        eqs.add(INCH_TO_MM[token])
        eqs.add("dn" + INCH_TO_MM[token])
        return eqs
    m_fen = re.fullmatch(r"(\d+)分", token)
    if m_fen:
        n = m_fen.group(1)
        if n in FEN_TO_MM_STR:
            mm = FEN_TO_MM_STR[n]
            eqs.add("dn" + mm)
            eqs.add(mm)
            if mm in MM_TO_INCH:
                eqs.add(MM_TO_INCH[mm])
        return eqs
    m_cun = re.fullmatch(r"(\d+)寸", token)
    if m_cun:
        n = m_cun.group(1)
        if n in CUN_INTEGER_TO_MM_STR:
            mm = CUN_INTEGER_TO_MM_STR[n]
            eqs.add("dn" + mm)
            eqs.add(mm)
            if mm in MM_TO_INCH:
                eqs.add(MM_TO_INCH[mm])
        return eqs
    return eqs


def _expand_token_with_synonyms_and_units(token: str, material: Optional[str] = None) -> set:
    synonyms = _get_synonym_words(token)
    expanded: set = set()
    for syn in synonyms:
        expanded |= _expand_unit_tokens(syn, material=material)
    return expanded


def _is_inch_token(token: str) -> bool:
    """是否为英寸规格 token（如 3/4"、1-1/4"、4"）。"""
    t = (token or "").strip()
    return bool(
        re.fullmatch(r"\d+(?:-\d+)?/\d+\"", t)
        or re.fullmatch(r"\d+(?:\.\d+)?\"", t)
    )


def _should_apply_inch_exact_priority(
    query_size_tokens: set[str],
    query_inch_tokens: set[str],
) -> bool:
    """
    「英寸精确命中优先」剪切：仅当查询里的英寸与 DN 数字在标准对照上一致时才启用。

    若用户混写互斥的公称径与英寸（例如 DN20 对应 3/4\"，却写了 1/2\"≈DN16），
    则不做剪切，否则仅写 dn20、库内无字面 1/2\" 的正确行会被误杀，而带 3/4\"x1/2\" 的管件会留下。
    """
    if not query_inch_tokens:
        return False
    dn_numbers: set[str] = set()
    for t in query_size_tokens:
        tl = (t or "").lower().strip()
        if tl.startswith("dn") and len(tl) > 2 and tl[2:].isdigit():
            dn_numbers.add(tl[2:])
    if not dn_numbers:
        return True
    for inch_t in query_inch_tokens:
        mm = INCH_TO_MM.get(inch_t)
        if mm is not None and mm not in dn_numbers:
            return False
    return True


def _split_tokens(text: str) -> List[str]:
    text = _normalize(text)
    tokens: List[str] = []
    # 先提取英寸分数规格，避免被后续纯数字提取拆成 3、4 这类噪声 token
    # 例：3/4"、1-1/4"
    for m in re.finditer(r"\d+(?:\s*-\s*\d+)?\s*/\s*\d+\s*[\"”″]?", text):
        raw = m.group()
        compact = re.sub(r"\s+", "", raw)
        compact = compact.replace("”", '"').replace("″", '"')
        if "/" in compact:
            if not compact.endswith('"'):
                compact = compact + '"'
            tokens.append(compact)
    text = re.sub(r"\d+(?:\s*-\s*\d+)?\s*/\s*\d+\s*[\"”″]?", " ", text)
    for m in re.finditer(r"\d+(?:\.\d+)?\s*[\"”″]", text):
        raw = m.group()
        compact = re.sub(r"\s+", "", raw)
        compact = compact.replace("”", '"').replace("″", '"')
        if not compact.endswith('"'):
            compact = compact + '"'
        tokens.append(compact)
    text = re.sub(r"\d+(?:\.\d+)?\s*[\"”″]", " ", text)
    # 「N分」口语（4分=DN15、6分=DN20）：整段提取，避免拆成孤立数字「4」导致规格过滤失败
    for m in re.finditer(r"\d+\s*分", text):
        tokens.append(re.sub(r"\s+", "", m.group()))
    text = re.sub(r"\d+\s*分", " ", text)
    # 「N寸」整数寸（如 2寸→DN50）
    for m in re.finditer(r"\d+\s*寸", text):
        tokens.append(re.sub(r"\s+", "", m.group()))
    text = re.sub(r"\d+\s*寸", " ", text)
    for m in re.finditer(r"dn\s*(\d+)", text, re.I):
        tokens.append("dn" + m.group(1))
        tokens.append(m.group(1))
    text = re.sub(r"dn\s*\d+", " ", text, flags=re.I)
    for m in re.finditer(r"\d+(?:\.\d+)?", text):
        tokens.append(m.group())
    text = re.sub(r"\d+(?:\.\d+)?", " ", text)
    for m in re.finditer(r"[\u4e00-\u9fff]+", text):
        tok = m.group()
        tokens.append(tok)
        # 长中文整段（如「度弯头带检查口」）在品名中常被括号/符号隔开，拆成单字以便「度」→「°」、「弯头」「带」「检查口」等能分别命中
        if len(tok) > 2:
            for c in tok:
                tokens.append(c)
    return list(dict.fromkeys(tokens))


def _expand_keyword_with_synonyms(keyword: str) -> List[str]:
    queries: set = {keyword}
    for syn in _SORTED_SYNONYMS:
        new_queries: set = set()
        for q in queries:
            if syn in q:
                for replacement in _SYNONYM_TO_GROUP[syn]:
                    new_queries.add(q.replace(syn, replacement))
        if new_queries:
            queries.update(new_queries)
    return list(queries)


def _safe_to_float(val: Any) -> Optional[float]:
    """将单元格值安全转为 float；非法或空值返回 None。"""
    if val is None:
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    if f != f or f in (float("inf"), float("-inf")):
        return None
    return f


def _normalize_code_for_match(val: Any) -> str:
    """物料编号用于匹配时的规范化：数字去掉尾部的 .0，避免 8010072480.0 与 8010072480 不匹配。"""
    if val is None or (hasattr(pd, "isna") and pd.isna(val)):
        return ""
    s = str(val).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except (TypeError, ValueError):
        return s


def normalize_price(raw_price: Any) -> float:
    """
    规范化用户输入的价格为 float（人民币金额）。

    规则（auto_fix 风格）：
    - 若本身是数字（int/float），直接返回 float 值。
    - 若是字符串：
      - 去除首尾空白、常见货币符号（¥/￥/元/RMB 等）及内部空格。
      - 将全角逗号替换为半角，移除千分位逗号。
      - 若存在多个小数点：保留最后一个点作为小数点，其余点视作分隔符移除（例如 "7.858.0" → "7858.0"）。
      - 清洗后字符串必须形如 `^[+-]?\\d+(\\.\\d+)?$`，否则视为格式不合法。
    - 解析失败或结果为 NaN/无穷大时抛出 ValueError。
    """
    # 已经是数字的情况，直接信任调用方
    if isinstance(raw_price, (int, float)) and not isinstance(raw_price, bool):
        value = float(raw_price)
        if value != value or value in (float("inf"), float("-inf")):
            raise ValueError(f"价格数值非法: {raw_price!r}")
        return value

    s = str(raw_price).strip()
    if not s:
        raise ValueError("价格不能为空。")

    # 去掉常见货币符号与全角变体
    for sym in ("¥", "￥", "元", "RMB", "rmb"):
        s = s.replace(sym, "")

    # 统一全角/半角逗号与点
    s = s.replace("，", ",").replace("．", ".")
    # 去掉空格
    s = s.replace(" ", "")

    # 先移除千分位逗号
    s = s.replace(",", "")

    # 处理多个小数点：保留最后一个，其余删除
    if s.count(".") > 1:
        last_dot = s.rfind(".")
        cleaned_chars = []
        for idx, ch in enumerate(s):
            if ch == "." and idx != last_dot:
                continue
            cleaned_chars.append(ch)
        s = "".join(cleaned_chars)

    # 允许前导正负号，其余必须是数字或至多一个小数点
    import re as _re

    if not _re.fullmatch(r"[+-]?\d+(\.\d+)?", s):
        raise ValueError(f"价格格式不合法: {raw_price!r}")

    try:
        value = float(s)
    except (TypeError, ValueError):
        raise ValueError(f"价格无法解析: {raw_price!r}")

    if value != value or value in (float("inf"), float("-inf")):
        raise ValueError(f"价格数值非法: {raw_price!r}")
    return value


def _contains_any(text: str, patterns: list[str]) -> bool:
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in patterns)


def _is_pressure_token(token: str) -> bool:
    text = _normalize(token or "")
    return bool(
        re.fullmatch(r"pn\s*[0-9]+(?:\.[0-9]+)?", text)
        or re.fullmatch(r"[0-9]+(?:\.[0-9]+)?\s*mpa", text)
    )


def _query_ceiling(norm_kw: str) -> bool:
    return _contains_any(
        norm_kw,
        [
            r"\bceiling\b",
            r"\bmain\s*hollow\b",
            r"\bstelldrat\b",
            r"\bsteel\s*drat\b",
            r"\bsoldays\b",
            r"\bdynabolt\b",
            r"\bmur\s+soldays\b",
        ],
    )


def _product_ceiling(product_text: str, product_type: str = "") -> bool:
    text = _normalize(f"{product_text} {product_type}")
    if re.search(r"\bceiling\b", _normalize(product_type or "")):
        return True
    return _contains_any(
        text,
        [
            r"\bceiling\b",
            r"\bmain\s*hollow\b",
            r"\bstelldrat\b",
            r"\bsteel\s*drat\b",
            r"\bsoldays\b",
            r"\bdynabolt\b",
        ],
    )


def _query_ceiling_category(norm_kw: str) -> str | None:
    if _contains_any(norm_kw, [r"\bmain\s*hollow\b", r"\bhollow\b"]):
        return "main_hollow"
    if _contains_any(norm_kw, [r"\bsteel\s*drat\b", r"\bstelldrat\b"]):
        return "stelldrat"
    if _contains_any(norm_kw, [r"\bdynabolt\b"]):
        return "dynabolt"
    if _contains_any(norm_kw, [r"\bmur\b"]):
        return "mur"
    if _contains_any(norm_kw, [r"\bmetal\s*ceiling\b", r"\bsnap\s*in\b", r"\bkeel\b"]):
        return "panel"
    if _contains_any(norm_kw, [r"\bhook\b"]):
        return "hook"
    return None


def _product_ceiling_category(product_text: str) -> str | None:
    text = _normalize(product_text or "")
    if _contains_any(text, [r"\bmain\s*hollow\b", r"\bhollow\b"]):
        return "main_hollow"
    if _contains_any(text, [r"\bstelldrat\b", r"\bsteel\s*drat\b"]):
        return "stelldrat"
    if _contains_any(text, [r"\bdynabolt\b"]):
        return "dynabolt"
    if _contains_any(text, [r"\bmur\b"]):
        return "mur"
    if _contains_any(text, [r"\bmetal\s*ceiling\b", r"\bsnap\s*in\b", r"\bkeel\b", r"\btriangle\b"]):
        return "panel"
    if _contains_any(text, [r"\bhook\b"]):
        return "hook"
    return None


def _query_rucika(norm_kw: str) -> bool:
    return bool(re.search(r"\brucika\b", norm_kw))


def _product_rucika_brand(product_text: str, product_type: str = "") -> bool:
    """RUCIKA 品牌产品线（RUCIKA STANDARD/JIS），不含仅后缀带 - RUCIKA 的联塑日标件。"""
    ptype = _normalize(product_type or "")
    return bool(re.search(r"\brucika\b", ptype))


def _looks_like_material_code(text: str) -> bool:
    s = (text or "").strip()
    if not s or " " in s:
        return False
    if re.fullmatch(r"\d{8,14}", s):
        return True
    return bool(re.fullmatch(r"[A-Za-z]{2,}(?:-[A-Za-z0-9]+)+", s))


def _query_material(norm_kw: str) -> str | None:
    if re.search(r"\bhdpe\b", norm_kw):
        return "hdpe"
    if re.search(r"\bppr\b", norm_kw):
        return "ppr"
    if re.search(r"\bpvc\b|\bupvc\b|\bcpvc\b", norm_kw):
        return "pvc"
    if re.search(r"\bpe\b", norm_kw):
        return "pe"
    return None


def _product_material(product_text: str, product_type: str = "") -> str | None:
    text = _normalize(f"{product_text} {product_type}")
    if re.search(r"\bppr\b", text):
        return "ppr"
    if re.search(r"\bhdpe\b", text):
        return "hdpe"
    if re.search(r"\bpe\b", text):
        return "pe"
    if re.search(r"\bpvc\b|\bupvc\b|\bcpvc\b|rucika|\baw\b|\bd\b", text):
        return "pvc"
    return None


def _query_usage(norm_kw: str) -> str | None:
    if _contains_any(norm_kw, [r"穿线", r"电线", r"线管", r"conduit", r"electrical"]):
        return "conduit"
    if _contains_any(norm_kw, [r"排水", r"排污", r"下水", r"drain"]):
        return "drain"
    if _contains_any(norm_kw, [r"给水", r"供水", r"自来水", r"water supply", r"pressure"]):
        return "water"
    return None


def _product_usage(product_text: str, code: str = "") -> str | None:
    text = _normalize(product_text)
    if code.startswith(("803002", "803004", "803005")):
        return "conduit"
    if _contains_any(text, [r"穿线", r"电线", r"线管", r"conduit", r"electrical"]):
        return "conduit"
    if _contains_any(text, [r"排水", r"排污", r"下水", r"drainage", r"drain", r"\bd\s*pipe\b"]):
        return "drain"
    if _contains_any(text, [r"给水", r"供水", r"自来水", r"water supply", r"\baw\b"]):
        return "water"
    return None


def _query_fitting(norm_kw: str) -> str | None:
    categories: list[tuple[str, list[str]]] = [
        ("glue", [r"胶水", r"胶粘剂", r"cement", r"glue"]),
        ("welder", [r"热熔器", r"焊接机", r"welding machine", r"welder"]),
        ("hose", [r"软管", r"\bhose\b", r"flexible pipe", r"flexible tube"]),
        ("triangle_valve", [r"三角阀", r"triangle valve"]),
        ("angle_valve", [r"角阀", r"angle valve"]),
        ("faucet", [r"水龙头", r"龙头", r"tap", r"faucet"]),
        ("valve", [r"阀", r"valve"]),
        ("tee", [r"三通", r"\btee\b"]),
        ("elbow", [r"弯头", r"弯管", r"弯", r"elbow"]),
        ("reducer", [r"变径", r"异径", r"大小头", r"reduc"]),
        ("cap", [r"管帽", r"堵头", r"端盖", r"\bcap\b", r"\bplug\b"]),
        ("coupling", [r"直接", r"直通", r"管箍", r"套筒", r"coupling", r"socket"]),
    ]
    for category, patterns in categories:
        if _contains_any(norm_kw, patterns):
            return category
    if _contains_any(norm_kw, [r"管件", r"fitting"]):
        return "fitting"
    if _contains_any(norm_kw, [r"管材", r"水管", r"管子", r"\bpipe\b"]):
        return "pipe"
    return None


def _product_fitting(product_text: str) -> str | None:
    text = _normalize(product_text)
    if _contains_any(text, [r"faucet\s*elbow", r"内螺纹弯头", r"丝扣弯头", r"螺纹弯头"]):
        return "elbow"
    categories: list[tuple[str, list[str]]] = [
        ("glue", [r"胶水", r"胶粘剂", r"cement", r"glue"]),
        ("welder", [r"热熔器", r"焊接机", r"welding machine", r"welder"]),
        ("hose", [r"软管", r"\bhose\b", r"flexible pipe", r"flexible tube"]),
        ("triangle_valve", [r"三角阀", r"triangle valve"]),
        ("angle_valve", [r"角阀", r"angle valve"]),
        ("faucet", [r"水龙头", r"龙头", r"tap", r"faucet"]),
        ("valve", [r"阀", r"valve"]),
        ("tee", [r"三通", r"\btee\b"]),
        ("elbow", [r"弯头", r"弯管", r"elbow"]),
        ("reducer", [r"变径", r"异径", r"大小头", r"reduc"]),
        ("cap", [r"管帽", r"堵头", r"端盖", r"\bcap\b", r"\bplug\b"]),
        ("coupling", [r"直接", r"直通", r"管箍", r"套筒", r"coupling", r"socket"]),
        ("pipe", [r"管材", r"水管", r"\bpipe\b", r"\btube\b"]),
    ]
    for category, patterns in categories:
        if _contains_any(text, patterns):
            return category
    return None


def _thread_gender(text: str) -> str | None:
    norm_text = _normalize(text)
    if _contains_any(norm_text, [r"内丝", r"内螺纹", r"female thread", r"\bfemale\b"]):
        return "female"
    if _is_fitting_drat_thread(norm_text):
        return "female"
    if _contains_any(norm_text, [r"丝扣", r"螺纹弯头", r"螺纹弯"]) and _query_fitting(norm_text) in {
        "elbow",
        "tee",
        "coupling",
        "reducer",
        "fitting",
    }:
        return "female"
    if _contains_any(norm_text, [r"外丝", r"外螺纹", r"male thread", r"\bmale\b"]):
        return "male"
    return None


def _extract_query_pressure(norm_kw: str) -> float | None:
    match = re.search(r"\bpn\s*([0-9]+(?:\.[0-9]+)?)\b", norm_kw)
    if match:
        # Inverse of _apply_pressure_expansion: PN数字 = MPa × 10（PN6→0.6, PN16→1.6）
        return float(match.group(1)) / 10
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*mpa\b", norm_kw)
    if match:
        return float(match.group(1))
    return None


def _canon_numeric_token(value: str) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value).strip().lower()
    if number.is_integer():
        return str(int(number))
    return f"{number:.4f}".rstrip("0").rstrip(".")


def _canon_size_component(value: str) -> str:
    token = _normalize(value or "")
    token = token.strip().strip('"＂')
    token = token.replace("inch", "").replace("in", "").strip()
    if token.startswith(("dn", "de", "od")):
        token = re.sub(r"^(?:dn|de|od)\s*", "", token)
    if re.fullmatch(r"\d+(?:\.\d+)?", token):
        return _canon_numeric_token(token)
    return token


def _sub_size_equivalents(value: str, thread_side: bool = False) -> set[str]:
    component = _canon_size_component(value)
    equivalents = set()
    if thread_side and component in THREAD_SIDE_TO_INCH:
        equivalents.add(THREAD_SIDE_TO_INCH[component].strip('"'))
        return equivalents
    equivalents.add(component)
    if component in MM_TO_INCH:
        equivalents.add(MM_TO_INCH[component].strip('"'))
    if component + '"' in INCH_TO_MM:
        equivalents.add(INCH_TO_MM[component + '"'])
    if component in INCH_TO_MM:
        equivalents.add(INCH_TO_MM[component])
    return equivalents


_SINGLE_DN_FITTINGS = frozenset({
    "coupling",
    "elbow",
    "cap",
    "tee",
    "valve",
    "triangle_valve",
    "angle_valve",
    "faucet",
})


def _has_thread_compound_context(norm_text: str) -> bool:
    if _thread_gender(norm_text):
        return True
    return _contains_any(norm_text, [r"\bfemale\b", r"\bmale\b", r"内丝", r"外丝", r"\bthread\b"])


def _is_dn_dn_reducer_compound(norm_text: str, sub_raw: str) -> bool:
    """PPR Reducing 40x32 style: both sides are metric DN, not inch thread sub."""
    if _has_thread_compound_context(norm_text):
        return False
    if _query_fitting(norm_text) == "reducer":
        return True
    if _contains_any(norm_text, [r"\breducing\b", r"大小头", r"异径套", r"异径直通"]):
        return True
    sub = _canon_size_component(sub_raw)
    return sub.isdigit() and int(sub) >= 20


def _compound_sub_thread_side(norm_text: str, prefix: str, sub_raw: str) -> bool:
    if prefix:
        return False
    if _is_dn_dn_reducer_compound(norm_text, sub_raw):
        return False
    return True


def _bare_nominal_dn_values(norm_text: str) -> set[str]:
    """Extract bare DN numbers from VANTSING-style fitting queries (e.g. Coupling 40)."""
    if _compound_specs_without_bare_scan(norm_text):
        return set()
    q_fit = _query_fitting(norm_text)
    if q_fit not in _SINGLE_DN_FITTINGS:
        return set()

    angle_nums = {_canon_numeric_token(str(a)) for a in _angle_values(norm_text)}
    values: set[str] = set()
    for match in re.finditer(r"\b([0-9]+(?:\.[0-9]+)?)\b", norm_text):
        raw = match.group(1)
        token = _canon_numeric_token(raw)
        if token in angle_nums:
            continue
        tail = norm_text[match.end() :]
        if re.match(r"\s*[x×*]", tail):
            continue
        head = norm_text[max(0, match.start() - 3) : match.start()]
        if re.search(r"pn\s*$", head, re.IGNORECASE):
            continue
        if re.match(r"\s*m(?![a-z])", tail):
            continue
        if token.isdigit() and 10 <= int(token) <= 400:
            values.add(token)
    return values


def _compound_specs_without_bare_scan(text: str) -> set[tuple[str, frozenset[str]]]:
    """Compound specs only (used by _bare_nominal_dn_values to avoid recursion)."""
    norm_text = _normalize(text)
    specs: set[tuple[str, frozenset[str]]] = set()
    side_pattern = r'([0-9]+(?:\.[0-9]+)?(?:\s*-\s*[0-9]+)?(?:/[0-9]+)?\s*(?:"|＂|in|inch)?)'
    for match in re.finditer(
        rf"\b(dn|de|od)?\s*([0-9]+(?:\.[0-9]+)?)\s*[x×*]\s*{side_pattern}",
        norm_text,
    ):
        prefix = (match.group(1) or "").strip()
        main = _canon_numeric_token(match.group(2))
        sub = _sub_size_equivalents(
            match.group(3),
            thread_side=_compound_sub_thread_side(norm_text, prefix, match.group(3)),
        )
        if main and sub:
            specs.add((main, frozenset(sub)))

    for match in re.finditer(r"\b([1-9][0-9])([1-9][0-9])(?=[\u4e00-\u9fff]|$)", norm_text):
        main = _canon_numeric_token(match.group(1))
        sub = _sub_size_equivalents(match.group(2), thread_side=True)
        if main and sub:
            specs.add((main, frozenset(sub)))
    return specs


def _compound_specs(text: str) -> set[tuple[str, frozenset[str]]]:
    norm_text = _normalize(text)
    specs: set[tuple[str, frozenset[str]]] = set()
    side_pattern = r'([0-9]+(?:\.[0-9]+)?(?:\s*-\s*[0-9]+)?(?:/[0-9]+)?\s*(?:"|＂|in|inch)?)'
    for match in re.finditer(
        rf"\b(dn|de|od)?\s*([0-9]+(?:\.[0-9]+)?)\s*[x×*]\s*{side_pattern}",
        norm_text,
    ):
        prefix = (match.group(1) or "").strip()
        main = _canon_numeric_token(match.group(2))
        sub = _sub_size_equivalents(
            match.group(3),
            thread_side=_compound_sub_thread_side(norm_text, prefix, match.group(3)),
        )
        if main and sub:
            specs.add((main, frozenset(sub)))

    # 口语紧凑写法，如 3220内丝三通 => dn32 x 1/2"。
    for match in re.finditer(r"\b([1-9][0-9])([1-9][0-9])(?=[\u4e00-\u9fff]|$)", norm_text):
        main = _canon_numeric_token(match.group(1))
        sub = _sub_size_equivalents(match.group(2), thread_side=True)
        if main and sub:
            specs.add((main, frozenset(sub)))
    return specs


def _diameter_values(text: str) -> set[str]:
    norm_text = _normalize(text)
    values: set[str] = set()
    for match in re.finditer(r"\b(?:dn|de|od|d)\s*([0-9]+(?:\.[0-9]+)?)(?=\b|x|×|\*)", norm_text):
        values.add(_canon_numeric_token(match.group(1)))
    for match in re.finditer(r"([0-9]+(?:\.[0-9]+)?)\s*mm\b", norm_text):
        values.add(_canon_numeric_token(match.group(1)))
    for match in re.finditer(r"\b([0-9]+(?:/[0-9]+)?|[0-9]+(?:\.[0-9]+)?)\s*(?:in|inch|英寸|寸|[\"＂])", norm_text):
        values.add(match.group(1).strip().lower())
    values |= _bare_nominal_dn_values(norm_text)
    return values


def _angle_values(text: str) -> set[int]:
    norm_text = _normalize(text)
    values: set[int] = set()
    for match in re.finditer(r"\b(22\.5|45|90)\s*(?:°|度|deg|degree)?", norm_text):
        values.add(int(float(match.group(1))))
    if _contains_any(norm_text, [r"半弯", r"45\s*度", r"45\s*°", r"45\s*deg"]):
        values.add(45)
    if _contains_any(norm_text, [r"全弯", r"直角弯", r"90\s*度", r"90\s*°", r"90\s*deg"]):
        values.add(90)
    return values


def _query_angle_preference(norm_kw: str) -> int | None:
    angles = _angle_values(norm_kw)
    if 45 in angles:
        return 45
    if 90 in angles:
        return 90
    if _query_fitting(norm_kw) == "elbow":
        return 90
    return None


def _product_pressure_values(product_text: str) -> set[float]:
    text = _normalize(product_text)
    values: set[float] = set()
    for match in re.finditer(r"\bpn\s*([0-9]+(?:\.[0-9]+)?)\b", text):
        value = float(match.group(1))
        values.add(value / 10 if value >= 10 else value)
    for match in re.finditer(r"([0-9]+(?:\.[0-9]+)?)\s*mpa\b", text):
        values.add(float(match.group(1)))
    return values


def _extract_query_length(norm_kw: str) -> float | None:
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(?:m(?![a-z])|米)\s*(?:/|每)?\s*(?:根|pcs|pc)?", norm_kw)
    if match:
        return float(match.group(1))
    return None


def _product_length(product_text: str) -> float | None:
    text = _normalize(product_text)
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(?:m(?![a-z])|米)\s*(?:/|每)?\s*(?:根|pcs|pc)?", text)
    if match:
        return float(match.group(1))
    return None


def _hard_filter_and_bonus(
    norm_kw: str,
    product_text: str,
    code: str = "",
    product_type: str = "",
) -> tuple[bool, float]:
    product_norm = _normalize(f"{product_text} {product_type}")
    bonus = 0.0

    if _query_ceiling(norm_kw):
        if not _product_ceiling(product_norm, product_type):
            return False, 0.0
        bonus += 0.2
        q_ceiling_cat = _query_ceiling_category(norm_kw)
        p_ceiling_cat = _product_ceiling_category(product_norm)
        if q_ceiling_cat and p_ceiling_cat and q_ceiling_cat != p_ceiling_cat:
            return False, 0.0
        if q_ceiling_cat and p_ceiling_cat and q_ceiling_cat == p_ceiling_cat:
            bonus += 0.2

    if _query_rucika(norm_kw):
        if not _product_rucika_brand(product_norm, product_type):
            return False, 0.0
        bonus += 0.25
    elif _product_rucika_brand(product_norm, product_type):
        return False, 0.0

    q_material = _query_material(norm_kw)
    p_material = _product_material(product_norm, product_type)
    if q_material and p_material and q_material != p_material:
        if not (q_material == "pe" and p_material == "hdpe"):
            return False, 0.0
    if q_material and p_material:
        bonus += 0.08

    q_usage = _query_usage(norm_kw)
    p_usage = _product_usage(product_norm, code)
    if q_usage and p_usage and q_usage != p_usage:
        return False, 0.0
    if q_usage and p_usage:
        bonus += 0.12

    q_fit = _query_fitting(norm_kw)
    p_fit = _product_fitting(product_norm)
    q_thread = _thread_gender(norm_kw)
    p_thread = _thread_gender(product_norm)
    non_pipe_categories = {
        "tee",
        "elbow",
        "reducer",
        "coupling",
        "valve",
        "triangle_valve",
        "angle_valve",
        "faucet",
        "cap",
        "glue",
        "hose",
        "welder",
    }
    exact_categories = {
        "tee",
        "elbow",
        "reducer",
        "coupling",
        "valve",
        "triangle_valve",
        "angle_valve",
        "faucet",
        "cap",
        "glue",
        "hose",
        "welder",
    }
    if q_fit == "pipe" and p_fit in non_pipe_categories:
        return False, 0.0
    if q_fit in exact_categories:
        if p_fit and p_fit != q_fit:
            threaded_faucet_elbow = (
                q_fit == "elbow"
                and q_thread == "female"
                and p_fit == "faucet"
                and p_thread == "female"
            )
            if not threaded_faucet_elbow:
                return False, 0.0
        if p_fit is None and q_fit != "glue":
            return False, 0.0
    if q_fit == "fitting" and p_fit in {"pipe", "faucet", "glue", "hose", "welder"}:
        return False, 0.0
    if q_fit and p_fit == q_fit:
        bonus += 0.18
    if q_material == "pvc" and q_usage is None and q_fit == "pipe":
        if p_usage == "drain":
            bonus += 0.14
        elif p_usage == "water":
            bonus -= 0.04

    q_compounds = _compound_specs(norm_kw)
    p_compounds = _compound_specs(product_norm)
    if q_compounds:
        if p_compounds:
            matched_compound = any(
                q_main == p_main and not q_subs.isdisjoint(p_subs)
                for q_main, q_subs in q_compounds
                for p_main, p_subs in p_compounds
            )
            if not matched_compound:
                return False, 0.0
            bonus += 0.18
        elif q_fit in {"tee", "elbow", "reducer", "coupling"}:
            return False, 0.0

    if q_fit == "tee" and not _contains_any(norm_kw, [r"异径", r"变径", r"大小头", r"reduc", r"[x×*]"]):
        if _contains_any(product_norm, [r"异径", r"变径", r"reduc", r"[x×*]"]):
            bonus -= 0.1
        else:
            bonus += 0.08

    if q_thread:
        if p_thread != q_thread:
            return False, 0.0
        bonus += 0.12

    q_angle = _query_angle_preference(norm_kw)
    p_angles = _angle_values(product_norm)
    if q_angle is not None and p_angles:
        if q_angle not in p_angles:
            if not (_query_fitting(norm_kw) == "elbow" and not _angle_values(norm_kw)):
                return False, 0.0
            bonus -= 0.08
        else:
            bonus += 0.08

    q_diameters = _diameter_values(norm_kw)
    p_diameters = _diameter_values(product_norm)
    if q_diameters and p_diameters:
        if q_diameters.isdisjoint(p_diameters):
            return False, 0.0
        bonus += 0.1
        for dn in q_diameters:
            if re.search(rf"\bdn\s*{re.escape(dn)}\b", product_norm):
                bonus += 0.12
                break

    q_pressure = _extract_query_pressure(norm_kw)
    p_pressures = _product_pressure_values(product_norm)
    if q_pressure is not None and p_pressures:
        if not any(abs(q_pressure - pressure) < 0.001 for pressure in p_pressures):
            return False, 0.0
        bonus += 0.1

    q_length = _extract_query_length(norm_kw)
    p_length = _product_length(product_norm)
    if q_length is not None and p_length is not None:
        if abs(q_length - p_length) > 0.001:
            return False, 0.0
        bonus += 0.06

    if _contains_any(norm_kw, [r"热水", r"hot water"]) and _contains_any(product_norm, [r"冷水", r"cold water"]):
        return False, 0.0
    if _contains_any(norm_kw, [r"冷水", r"cold water"]) and _contains_any(product_norm, [r"热水", r"hot water"]):
        return False, 0.0

    if q_material in {"pe", "hdpe"}:
        wants_electrofusion = _contains_any(norm_kw, [r"电熔", r"electrofusion"])
        wants_butt = _contains_any(norm_kw, [r"对接", r"butt", r"welding"])
        is_electrofusion = _contains_any(product_norm, [r"电熔", r"electrofusion"])
        if wants_electrofusion and not is_electrofusion:
            return False, 0.0
        if not wants_electrofusion and is_electrofusion:
            return False, 0.0
        if wants_butt and is_electrofusion:
            return False, 0.0
        if wants_electrofusion:
            bonus += 0.14

    if q_fit == "glue" and not _contains_any(product_norm, [r"胶水", r"胶粘剂", r"cement", r"glue"]):
        return False, 0.0

    return True, bonus


def search_fuzzy(
    df: pd.DataFrame,
    keyword: str,
    field: str = "Describrition",
) -> List[tuple[dict[str, Any], float]]:
    """
    DataBase-style fuzzy search.
    Returns [(row_dict, score), ...] sorted by score desc.
    row_dict: {code, matched_name, unit_price}

    优化点：
    - 使用 load_wanding_df 预计算的 norm_text / spec_tokens 列，消除每行的 regex 开销
    - q_eq 提出行循环，每次查询只计算一次
    - set 交集（q_eq & product_specs）替代 any(eq in set for eq in set)
    - 单字 token 权重 _SINGLE_CHAR_WEIGHT（0.5），避免单字命中过度拉高得分
    """
    results: dict = {}
    has_precomputed = "norm_text" in df.columns and "spec_tokens" in df.columns

    for kw in _expand_keyword_with_synonyms(keyword.strip()):
        norm_kw = _normalize(kw)
        chinese_tokens = _split_tokens(norm_kw)
        material_tokens = re.findall(r"pvc|ppr|pe|hdpe", norm_kw)
        query_size_tokens = {t for t in chinese_tokens if re.search(r"\d", t) and not t.endswith("°")}
        query_inch_tokens = {t for t in query_size_tokens if _is_inch_token(t)}
        query_text_tokens = {
            t for t in chinese_tokens if not (re.search(r"\d", t) and not t.endswith("°"))
        }
        query_size_tokens = {t for t in query_size_tokens if not _is_pressure_token(t)}
        query_text_tokens = {t for t in query_text_tokens if not _is_pressure_token(t)}
        query_pressure = _extract_query_pressure(norm_kw)
        if query_pressure is not None:
            pressure_tokens = {
                _canon_numeric_token(str(query_pressure)),
                _canon_numeric_token(str(query_pressure * 10)),
            }
            query_size_tokens = {
                t for t in query_size_tokens if _canon_numeric_token(t) not in pressure_tokens
            }
        query_material = material_tokens[0] if material_tokens else None

        # q_eq 提出行循环：只依赖 q_spec + query_material，与当前行无关
        spec_equivs: dict[str, frozenset] = {
            q_spec: _expand_token_with_synonyms_and_units(q_spec, material=query_material)
            for q_spec in query_size_tokens
        }

        # 按单字/多字分类，单字在分母中按 _SINGLE_CHAR_WEIGHT 计入
        multi_text = {t for t in query_text_tokens if len(t) > 1}
        single_text = {t for t in query_text_tokens if len(t) == 1}
        total_weight = (
            len(query_size_tokens)
            + len(multi_text)
            + len(single_text) * _SINGLE_CHAR_WEIGHT
        )

        iter_rows: list[tuple[dict[str, Any], float, int]] = []
        for row in df.itertuples(index=False):
            row_id = getattr(row, "Material", getattr(row, "Describrition", str(row)))
            raw_text = str(getattr(row, field, ""))
            code_text = str(getattr(row, "Material", "") or "")
            product_type_text = str(getattr(row, "Product_Type", "") or "")
            product_text_for_rules = " ".join(
                [
                    raw_text,
                    str(getattr(row, "Describrition_English", "") or ""),
                    product_type_text,
                ]
            )

            # 使用预计算列，fallback 到实时计算（兼容未预计算的 df）
            if has_precomputed:
                normalized_text: str = row.norm_text
                product_specs: frozenset = row.spec_tokens
            else:
                normalized_text = _normalize(raw_text)
                product_specs = frozenset(
                    t for t in _split_tokens(raw_text) if re.search(r"\d", t)
                )

            keep_candidate, compat_bonus = _hard_filter_and_bonus(
                norm_kw,
                product_text_for_rules,
                code_text,
                product_type_text,
            )
            if not keep_candidate:
                continue

            if material_tokens and not all(m.lower() in normalized_text for m in material_tokens):
                if _query_material(norm_kw) != _product_material(product_text_for_rules, product_type_text):
                    continue

            if _query_fitting(norm_kw) == "glue" and not query_size_tokens:
                score = compat_bonus or 0.1
                row_dict = {
                    "code": code_text,
                    "matched_name": raw_text,
                }
                if hasattr(row, "unit_price"):
                    row_dict["unit_price"] = getattr(row, "unit_price", 0.0)
                if hasattr(row, "Product_Type"):
                    row_dict["Product_Type"] = product_type_text.strip()
                supplier = str(getattr(row, "supplier", "") or getattr(row, "Supplier", "") or "").strip()
                if supplier:
                    row_dict["supplier"] = supplier
                desc_en = str(getattr(row, "Describrition_English", "") or "").strip()
                if desc_en:
                    row_dict["description_english"] = desc_en
                iter_rows.append((row_dict, score, 0))
                continue

            # set 交集替代 any(eq in product_specs for eq in q_eq)
            size_hits = sum(1 for q_eq in spec_equivs.values() if q_eq & product_specs)
            if query_size_tokens and size_hits == 0:
                continue
            inch_exact_hits = sum(1 for t in query_inch_tokens if t in product_specs)

            # 多字命中（权重 1.0）+ 单字命中（权重 _SINGLE_CHAR_WEIGHT）
            def _text_match(t: str) -> bool:
                return t.lower() in normalized_text or (t == "度" and "°" in normalized_text)

            multi_hits = sum(1 for t in multi_text if _text_match(t))
            single_hits = sum(1 for t in single_text if _text_match(t))
            # 过滤用原始命中数（单字也算），得分用加权值
            if query_text_tokens and (multi_hits + single_hits) == 0:
                continue

            hit_weight = size_hits + multi_hits + single_hits * _SINGLE_CHAR_WEIGHT
            score = hit_weight / total_weight if total_weight > 0 else 0.0
            score += compat_bonus

            if score > 0 and (row_id not in results or score > results[row_id][1]):
                row_dict: dict[str, Any] = {
                    "code": code_text,
                    "matched_name": raw_text,
                }
                if hasattr(row, "unit_price"):
                    row_dict["unit_price"] = getattr(row, "unit_price", 0.0)
                if hasattr(row, "Product_Type"):
                    row_dict["Product_Type"] = product_type_text.strip()
                supplier = str(getattr(row, "supplier", "") or getattr(row, "Supplier", "") or "").strip()
                if supplier:
                    row_dict["supplier"] = supplier
                desc_en = str(getattr(row, "Describrition_English", "") or "").strip()
                if desc_en:
                    row_dict["description_english"] = desc_en
                iter_rows.append((row_dict, score, inch_exact_hits))

        # 英寸优先：若查询里显式给了英寸，且存在英寸精确命中的候选，
        # 则只保留英寸精确命中，避免被 dn 等价扩展引入跨体系误匹配。
        # （当 DN 与英寸互斥时跳过，见 _should_apply_inch_exact_priority）
        if (
            _should_apply_inch_exact_priority(query_size_tokens, query_inch_tokens)
            and query_inch_tokens
            and any(inch_hits > 0 for _, _, inch_hits in iter_rows)
        ):
            iter_rows = [r for r in iter_rows if r[2] > 0]

        for row_dict, score, _inch_hits in iter_rows:
            row_id = row_dict.get("code") or row_dict.get("matched_name")
            if row_id not in results or score > results[row_id][1]:
                results[row_id] = (row_dict, score)

    out = list(results.values())
    out.sort(key=lambda x: x[1], reverse=True)
    return out


def _load_one_sheet(ws, price_col: int) -> list[dict]:
    """从已打开的 worksheet 读出一张表的行（Material, Describrition, Describrition_English, Product_Type, unit_price）。需覆盖 E 档列(0-based 18)，故 max_col=20。"""
    rows = []
    for row in ws.iter_rows(max_col=20):
        cells = [getattr(c, "value", None) for c in row]
        if len(cells) > 2 and cells[2]:
            up = 0.0
            if len(cells) > price_col and cells[price_col] is not None:
                try:
                    up = float(cells[price_col])
                except (ValueError, TypeError):
                    pass
            rows.append({
                "Material": str(cells[1] or "").strip(),
                "Describrition": str(cells[2] or "").strip(),
                "Describrition_English": str(cells[3] or "").strip() if len(cells) > 3 else "",
                "Product_Type": str(cells[4] or "").strip() if len(cells) > 4 else "",
                "unit_price": up,
            })
    return rows


def _new_price_field_for_level(level: str) -> str:
    lu = _normalize_price_level(level)
    if lu == "FACTORY_INC_TAX":
        return "factory_inc_tax"
    if lu == "FACTORY_EXC_TAX":
        return "factory_exc_tax"
    if lu == "PURCHASE_EXC_TAX":
        return "purchase_exc_tax"
    if lu == "LOCAL_EXC_TAX":
        return "local_exc_tax"
    if lu == "LOCAL_INC_TAX":
        return "local_inc_tax"
    if lu == "RUCIKA_PRICELIST_EXC":
        return "rucika_pricelist_exc_vat11"
    if lu == "RUCIKA_PRICELIST_INC":
        return "rucika_pricelist_inc_vat11"
    if lu == "RUCIKA_QUOTE_1":
        return "rucika_quote_price_1"
    if lu == "RUCIKA_QUOTE_2":
        return "rucika_quote_price_2"
    if lu == "PE_NOMINAL":
        return "pe_nominal_price"
    if lu == "PE_FACTORY":
        return "pe_factory_price"
    if lu.startswith("A_") or lu == "A":
        return "price_a"
    if lu.startswith("B_") or lu == "B":
        return "price_b"
    if lu.startswith("C_") or lu == "C":
        return "price_c"
    if lu in ("D_LOW", "D_WHOLESALE"):
        return "price_d_low"
    if lu.startswith("D_") or lu == "D":
        return "price_d"
    if lu.startswith("E_") or lu == "E":
        return "price_e"
    return "price_b"


def _fallback_price_fields_for_level(level: str) -> list[str]:
    """Price columns to try in order. Extended tiers avoid cross-tier fallback to B."""
    lu = _normalize_price_level(level)
    field = _new_price_field_for_level(lu)
    extended: dict[str, list[str]] = {
        "LOCAL_EXC_TAX": ["local_exc_tax", "local_inc_tax"],
        "LOCAL_INC_TAX": ["local_inc_tax", "local_exc_tax"],
        "RUCIKA_PRICELIST_EXC": [
            "rucika_pricelist_exc_vat11",
            "factory_exc_tax",
            "rucika_quote_price_1",
        ],
        "RUCIKA_PRICELIST_INC": [
            "rucika_pricelist_inc_vat11",
            "factory_inc_tax",
            "rucika_quote_price_1",
        ],
        "RUCIKA_QUOTE_1": ["rucika_quote_price_1", "price_b"],
        "RUCIKA_QUOTE_2": ["rucika_quote_price_2", "price_d"],
        "PE_NOMINAL": ["pe_nominal_price", "pe_factory_price", "price_b"],
        "PE_FACTORY": [
            "pe_factory_price",
            "factory_exc_tax",
            "purchase_exc_tax",
            "price_b",
        ],
    }
    if lu in extended:
        seen: set[str] = set()
        ordered: list[str] = []
        for name in extended[lu]:
            if name not in seen:
                seen.add(name)
                ordered.append(name)
        return ordered
    chain = [
        field,
        "price_b",
        "factory_exc_tax",
        "purchase_exc_tax",
        "factory_inc_tax",
        "pe_factory_price",
        "rucika_quote_price_1",
    ]
    ordered: list[str] = []
    seen: set[str] = set()
    for name in chain:
        if name not in seen:
            seen.add(name)
            ordered.append(name)
    return ordered


def _coerce_bool(value: Any, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _safe_float_or_zero(value: Any) -> float:
    parsed = _safe_to_float(value)
    return float(parsed) if parsed is not None else 0.0


def _load_new_price_library_sheet(ws, customer_level: str) -> list[dict]:
    """Load unified price_library sheet and adapt it to legacy matcher columns."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v or "").strip() for v in rows[0]]
    header_idx = {name: idx for idx, name in enumerate(headers) if name}
    required = {"material", "description"}
    if not required.issubset(header_idx):
        return []
    out: list[dict] = []
    fallback_price_fields = _fallback_price_fields_for_level(customer_level)

    def cell(row: tuple[Any, ...], name: str) -> Any:
        idx = header_idx.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in rows[1:]:
        material = str(cell(row, "material") or "").strip()
        desc = str(cell(row, "description") or "").strip()
        if not material or not desc:
            continue
        if not _coerce_bool(cell(row, "is_preferred_price"), default=True):
            continue

        desc_cn = str(cell(row, "description_cn") or "").strip()
        desc_en = str(cell(row, "description_english") or "").strip()
        display_desc = desc_cn or desc
        unit_price = 0.0
        for field_name in fallback_price_fields:
            unit_price = _safe_float_or_zero(cell(row, field_name))
            if unit_price:
                break
        search_text = " ".join(
            part for part in [display_desc, desc, desc_cn, desc_en] if part
        )
        out.append(
            {
                "Material": material,
                "Describrition": display_desc,
                "Describrition_Raw": desc,
                "Describrition_CN": desc_cn,
                "Describrition_English": desc_en,
                "Product_Type": str(cell(row, "product_type") or "").strip(),
                "supplier": str(cell(row, "supplier") or "").strip(),
                "unit_price": unit_price,
                "source_file": str(cell(row, "source_file") or "").strip(),
                "source_sheet": str(cell(row, "source_sheet") or "").strip(),
                "source_row": cell(row, "source_row"),
                "search_text": search_text,
                "is_preferred_price": True,
            }
        )
    return out


def load_wanding_df(
    path: str | Path,
    sheet_name: str = "管材",
    customer_level: str = "B",
) -> pd.DataFrame:
    """Load 万鼎 price library as DataFrame. 默认加载「管材」+「国标管件」两个 sheet 并合并，以便匹配带检查口弯头、管帽等国标管件。"""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl 未安装，万鼎模糊匹配不可用")
        return pd.DataFrame()

    p = Path(path)
    if p.is_absolute() and p.exists():
        pass
    elif not p.is_absolute():
        root = Path(__file__).resolve().parent.parent.parent
        p = root / p
    if not p.exists():
        logger.warning("万鼎价格库不存在: %s", p)
        return pd.DataFrame()

    level = _normalize_price_level(customer_level)
    price_col = PRICE_COLS.get(level, PRICE_COLS["B"])

    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        all_rows: list[dict] = []
        if "price_library" in wb.sheetnames:
            all_rows.extend(_load_new_price_library_sheet(wb["price_library"], level))
        else:
            # 先加载管材
            ws_guan = wb["管材"] if "管材" in wb.sheetnames else (wb.active or wb[wb.sheetnames[0]])
            all_rows.extend(_load_one_sheet(ws_guan, price_col))
            # 若存在国标管件 sheet，一并加载（8020020643 带检查口弯头、8020020205 管帽等在此表）
            if "国标管件" in wb.sheetnames:
                all_rows.extend(_load_one_sheet(wb["国标管件"], price_col))
        wb.close()
        df = pd.DataFrame(all_rows)
        if not df.empty:
            # 预计算 normalized text 和规格 token 集，避免 search_fuzzy 每行重算
            search_series = df["search_text"] if "search_text" in df.columns else df["Describrition"]
            df["norm_text"] = search_series.apply(_normalize)
            df["spec_tokens"] = search_series.apply(
                lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
            )
        return df
    except Exception as e:
        logger.warning("加载万鼎价格库失败: %s", e)
        return pd.DataFrame()


def invalidate_wanding_cache() -> None:
    """清除万鼎 DataFrame 缓存（admin 更新 Neon 数据后调用）。"""
    with _df_cache_lock:
        _df_cache.clear()
    with _full_df_lock:
        _full_df_cache.clear()
    logger.info("wanding_fuzzy_matcher: DataFrame caches cleared")


def _level_to_db_price_field(level: str) -> Optional[str]:
    """A/B/C/D 档 → DB 列名；出厂/采购/E 档等返回 None（走本地 xlsx）。"""
    lu = level.upper()
    if "FACTORY" in lu or "PURCHASE" in lu:
        return None
    if lu.startswith("E_") or lu == "E":
        return None
    if lu.startswith("A_") or lu == "A":
        return "price_a"
    if lu.startswith("B_") or lu == "B":
        return "price_b"
    if lu.startswith("C_") or lu == "C":
        return "price_c"
    if lu.startswith("D_") or lu in ("D", "D_LOW", "D_NOADJ", "D_WHOLESALE"):
        return "price_d"
    return None


def _find_col(columns: list[dict], *keywords: str) -> str | None:
    """从 columns JSON 找第一个 name（lower）同时包含所有 keyword（lower）的列。"""
    for col in columns:
        name = (col.get("name") or "").lower()
        if all(kw.lower() in name for kw in keywords):
            return col["name"]
    return None


def _try_load_from_custom_library(level: str) -> Optional[pd.DataFrame]:
    """当 price_library 固定表为空时，从名含"万鼎"/"价格库"的自定义库拉数据。"""
    try:
        from admin import repository
        from inventory.config import config

        libs = repository.list_libraries()
        patterns = config.PRICE_LIB_NAME_PATTERNS
        matched = [lib for lib in libs if any(p in (lib.get("name") or "") for p in patterns)]
        if not matched:
            return None

        lib = max(matched, key=lambda x: x["id"])
        table_name = lib.get("table_name") or ""
        columns = lib.get("columns") or []
        if not table_name or not columns:
            return None

        col_material = _find_col(columns, config.PRICE_LIB_COL_MATERIAL_KW)
        col_desc = _find_col(columns, config.PRICE_LIB_COL_DESC_KW)
        field = _level_to_db_price_field(level)
        level_kw_map = {
            "price_a": config.PRICE_LIB_COL_PRICE_A_KW,
            "price_b": config.PRICE_LIB_COL_PRICE_B_KW,
            "price_c": config.PRICE_LIB_COL_PRICE_C_KW,
            "price_d": config.PRICE_LIB_COL_PRICE_D_KW,
        }
        price_kw = level_kw_map.get(field or "")
        col_price = _find_col(columns, *price_kw) if price_kw else None

        if not col_desc:
            logger.warning("自定义价格库找不到描述列 (keywords=%s)", config.PRICE_LIB_COL_DESC_KW)
            return None

        rows = repository.fetch_all_library_rows(table_name)
        if not rows:
            return None

        records = []
        for r in rows:
            material = str(r.get(col_material) or "" if col_material else "").strip()
            desc = str(r.get(col_desc) or "").strip()
            if not desc:
                continue
            up_f = 0.0
            if col_price:
                try:
                    up_f = float(r.get(col_price) or 0)
                except (TypeError, ValueError):
                    up_f = 0.0
            records.append({"Material": material, "Describrition": desc, "unit_price": up_f})

        if not records:
            return None
        df = pd.DataFrame(records)
        df["norm_text"] = df["Describrition"].apply(_normalize)
        df["spec_tokens"] = df["Describrition"].apply(
            lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
        )
        logger.info(
            "wanding_fuzzy_matcher: loaded %d rows from custom library '%s' (level=%s)",
            len(df), lib.get("name"), level,
        )
        return df
    except Exception as e:
        logger.warning("_try_load_from_custom_library 失败: %s", e)
        return None


def _try_load_from_db(level: str) -> Optional[pd.DataFrame]:
    """从 Neon admin 缓存行构建 DataFrame；无数据或不可用则返回 None（fallback xlsx）。"""
    try:
        from admin.cache import get_price_library_rows

        rows = get_price_library_rows()
        if not rows:
            return _try_load_from_custom_library(level)
        field = _level_to_db_price_field(level)
        if field is None:
            return None
        records = []
        for r in rows:
            if not _coerce_bool(r.get("is_preferred_price"), default=True):
                continue
            up = r.get(field)
            try:
                up_f = float(up) if up is not None else 0.0
            except (TypeError, ValueError):
                up_f = 0.0
            desc = str(r.get("description") or "").strip()
            desc_cn = str(r.get("description_cn") or "").strip()
            desc_en = str(r.get("description_english") or "").strip()
            display_desc = desc_cn or desc
            records.append(
                {
                    "Material": str(r.get("material") or "").strip(),
                    "Describrition": display_desc,
                    "Describrition_Raw": desc,
                    "Describrition_CN": desc_cn,
                    "Describrition_English": desc_en,
                    "Product_Type": str(r.get("product_type") or "").strip(),
                    "unit_price": up_f,
                    "search_text": " ".join(part for part in [display_desc, desc, desc_cn, desc_en] if part),
                }
            )
        df = pd.DataFrame(records)
        if df.empty:
            return None
        search_series = df["search_text"] if "search_text" in df.columns else df["Describrition"]
        df["norm_text"] = search_series.apply(_normalize)
        df["spec_tokens"] = search_series.apply(
            lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
        )
        logger.info("wanding_fuzzy_matcher: loaded %d rows from DB (level=%s)", len(df), level)
        return df
    except Exception as e:
        logger.warning("_try_load_from_db 失败，将 fallback 读 xlsx: %s", e)
        return None


# 缓存 DataFrame，按 path:level 隔离
_df_cache: dict[str, pd.DataFrame] = {}
_df_cache_lock = threading.Lock()

# Remote price source metadata keyed by normalised level.
# Written by _try_load_from_org_remote; read by match_fuzzy / match_fuzzy_candidates.
_remote_price_meta: dict[str, dict[str, Any]] = {}
_remote_meta_lock = threading.Lock()


def get_remote_price_meta(level: str) -> dict[str, Any]:
    """Return the latest remote price metadata for a given normalised level.

    Returns an empty dict if no remote data has been loaded yet.
    """
    with _remote_meta_lock:
        return dict(_remote_price_meta.get(level, {}))


def _try_load_from_org_remote(level: str) -> Optional[pd.DataFrame]:
    """Attempt to build a price DataFrame from the org remote client.

    Source priority (handled inside org_price_client.get_price_data):
        1. Org API   (stale=False)
        2. LKG snapshot  (stale=True)
        3. Bundled seed  (stale=True)

    Stores source/stale metadata in ``_remote_price_meta[level]`` so that
    callers can attach it to tool results.

    Returns None on any failure so the caller falls through to Neon DB / xlsx.
    """
    try:
        from admin.org_price_client import get_price_data
        from inventory.price_loader import load_price_dataframe

        price_data = get_price_data()
        if not price_data.get("products"):
            return None

        df, meta = load_price_dataframe(price_data, customer_level=level)
        if df is None or (hasattr(df, "empty") and df.empty):
            return None

        with _remote_meta_lock:
            _remote_price_meta[level] = meta

        logger.info(
            "wanding_fuzzy_matcher: remote price data rows=%d source=%s stale=%s",
            len(df),
            meta.get("price_source"),
            meta.get("price_stale"),
        )
        return df
    except Exception as e:
        logger.warning("_try_load_from_org_remote 失败: %s", e)
        return None


def _get_cached_df(path, customer_level: str, *, try_remote: bool = False) -> pd.DataFrame:
    """线程安全地获取 DataFrame。

    Source precedence (when try_remote=True):
        0. Org remote (org_api → lkg_snapshot → bundled_seed)
        1. Neon DB
        2. Local xlsx

    When ``try_remote=False`` (explicit price_library_path supplied by caller),
    the remote client is skipped for backward compatibility with smoke/E2E tests.
    """
    level = _normalize_price_level(customer_level)
    cache_key = f"{path}:{level}"
    with _df_cache_lock:
        if cache_key not in _df_cache:
            if try_remote:
                df_remote = _try_load_from_org_remote(level)
                if df_remote is not None and not df_remote.empty:
                    _df_cache[cache_key] = df_remote
                    return _df_cache[cache_key]
            df_db = _try_load_from_db(level)
            if df_db is not None and not df_db.empty:
                _df_cache[cache_key] = df_db
            else:
                df_local = load_wanding_df(path, customer_level=level)
                if df_local.empty:
                    logger.error(
                        "price library missing or empty (legacy fallback disabled): %s",
                        path,
                    )
                _df_cache[cache_key] = df_local
        return _df_cache[cache_key]


# --------- 利润率查询（按 code / 完整名称 + 价格）---------

_full_df_cache: dict[str, pd.DataFrame] = {}
_full_df_lock = threading.Lock()


def _load_full_price_df(path: str | Path) -> pd.DataFrame:
    """加载完整万鼎价格库 DataFrame（包含所有价格与利润率列），供利润率查询使用。"""
    p = Path(path)
    if not p.is_absolute():
        root = Path(__file__).resolve().parent.parent.parent
        p = root / p
    cache_key = str(p.resolve()) if p.exists() else str(p)
    if cache_key in _full_df_cache:
        return _full_df_cache[cache_key]
    with _full_df_lock:
        if cache_key in _full_df_cache:
            return _full_df_cache[cache_key]
        if not p.exists():
            logger.warning("万鼎价格库不存在: %s", p)
            _full_df_cache[cache_key] = pd.DataFrame()
            return _full_df_cache[cache_key]
        try:
            # 同时加载「管材」与「国标管件」两个 sheet 并合并
            sheets = pd.read_excel(p, sheet_name=None)
            if "price_library" in sheets:
                df_new = sheets["price_library"]
                if "is_preferred_price" in df_new.columns:
                    df_new = df_new[df_new["is_preferred_price"].apply(lambda v: _coerce_bool(v, default=True))]
                _full_df_cache[cache_key] = df_new
                return _full_df_cache[cache_key]
            frames: list[pd.DataFrame] = []
            for name, df in sheets.items():
                if name in ("管材", "国标管件") or not frames:
                    frames.append(df)
            df_all = pd.concat(frames, ignore_index=True)
            _full_df_cache[cache_key] = df_all
            return _full_df_cache[cache_key]
        except Exception as e:
            logger.warning("加载完整万鼎价格库失败: %s", e)
            _full_df_cache[cache_key] = pd.DataFrame()
            return _full_df_cache[cache_key]


def _compute_profit_for_price(row: pd.Series, price: float) -> dict[str, Any]:
    """给定一行价格库记录与用户价，计算匹配档位及其利润率，并返回所有档位价格+利润率。

    行为：
    - 仅当某档位价格与给定 price 在容差内相等（绝对误差 ≤ _tolerance，_tolerance = max(0.01, |target|×1e-5)）时，
      才设置 matched_* 字段，避免浮点/四舍五入导致“有记录但匹配不到档位”。
    - 若没有任何档位价格在容差内与 price 相等，则 matched_price_level/matched_price/matched_profit 保持为 None，
      不再回退到“距离最小”的近似档位。
    """
    all_levels: list[dict[str, Any]] = []
    if any(name in row.index for name in ("price_a", "price_b", "price_c", "price_d", "price_d_low", "price_e")):
        named_levels = [
            ("A_QUOTE", "price_a", "profit_a"),
            ("B_QUOTE", "price_b", "profit_b"),
            ("C_QUOTE", "price_c", "profit_c"),
            ("D_QUOTE", "price_d", "profit_d"),
            ("D_LOW", "price_d_low", "profit_d_low"),
            ("E_QUOTE", "price_e", "profit_e"),
            ("LOCAL_EXC_TAX", "local_exc_tax", "local_profit"),
            ("LOCAL_INC_TAX", "local_inc_tax", "local_profit"),
            ("RUCIKA_QUOTE_1", "rucika_quote_price_1", "rucika_quote_profit_1"),
            ("RUCIKA_QUOTE_2", "rucika_quote_price_2", "rucika_quote_profit_2"),
        ]
        for level, price_field, profit_field in named_levels:
            price_val = _safe_to_float(row.get(price_field))
            if price_val is None or price_val == 0:
                continue
            all_levels.append(
                {
                    "level": level,
                    "price": price_val,
                    "profit": _safe_to_float(row.get(profit_field)),
                    "level_display": PRICE_LEVEL_DISPLAY_NAMES.get(level, level),
                }
            )
        matched_level = None
        matched_price = None
        matched_profit = None
        if all_levels:
            target = float(price)
            _tolerance = max(0.01, abs(target) * 1e-5)
            exact = [entry for entry in all_levels if abs(entry["price"] - target) <= _tolerance]
            if exact:
                best = exact[0]
                matched_level = best["level"]
                matched_price = best["price"]
                matched_profit = best["profit"]
        return {
            "code": str(row.get("material") or row.get("Material") or row.get("code") or "").strip(),
            "name": str(row.get("description_cn") or row.get("description") or row.get("Describrition") or row.get("Description") or "").strip(),
            "matched_price_level": matched_level,
            "matched_price": matched_price,
            "matched_profit": matched_profit,
            "all_levels": all_levels,
        }

    for level in ("A_QUOTE", "B_QUOTE", "C_QUOTE", "D_QUOTE", "D_LOW", "E_QUOTE"):
        price_col = PRICE_COLS.get(level)
        profit_col = PROFIT_COLS.get(level)
        if price_col is None:
            continue
        price_val = None
        if len(row) > price_col:
            price_val = _safe_to_float(row.iloc[price_col])
        if price_val is None or price_val == 0:
            continue
        profit_val = None
        if profit_col is not None and len(row) > profit_col:
            profit_val = _safe_to_float(row.iloc[profit_col])
        all_levels.append(
            {
                "level": level,
                "price": price_val,
                "profit": profit_val,
                "level_display": PRICE_LEVEL_DISPLAY_NAMES.get(level, level),
            }
        )
    matched_level = None
    matched_price = None
    matched_profit = None
    if all_levels:
        target = float(price)
        # 档位匹配：允许极小误差（浮点/四舍五入），避免库中 21810.0 与用户 21810 或 21809.99 判为不匹配
        _tolerance = max(0.01, abs(target) * 1e-5)
        exact = [entry for entry in all_levels if abs(entry["price"] - target) <= _tolerance]
        if exact:
            best = exact[0]
            matched_level = best["level"]
            matched_price = best["price"]
            matched_profit = best["profit"]
    return {
        "code": str(row.get("Material") or row.get("code") or "").strip(),
        "name": str(row.get("Describrition") or row.get("Description") or "").strip(),
        "matched_price_level": matched_level,
        "matched_price": matched_price,
        "matched_profit": matched_profit,
        "all_levels": all_levels,
    }


def get_profit_rows_by_code(code: str, price: float, path: str | Path) -> list[dict[str, Any]]:
    """按 Material code 精确过滤价格库，并为每行计算与给定价格对应的利润率。"""
    df = _load_full_price_df(path)
    if df.empty:
        return []
    code_norm = _normalize_code_for_match(code)
    if "material" in df.columns:
        mask = df["material"].apply(lambda v: _normalize_code_for_match(v) == code_norm)
        rows = df[mask]
    elif "Material" in df.columns:
        # 统一规范化：库中可能是 8010072480.0，用户传 8010072480，需一致
        mask = df["Material"].apply(lambda v: _normalize_code_for_match(v) == code_norm)
        rows = df[mask]
    else:
        rows = pd.DataFrame()
    return [_compute_profit_for_price(row, price) for _, row in rows.iterrows()]


def get_profit_rows_by_name(name: str, price: float, path: str | Path) -> list[dict[str, Any]]:
    """按完整中文名称过滤价格库，并为每行计算与给定价格对应的利润率。"""
    df = _load_full_price_df(path)
    if df.empty:
        return []
    name_norm = _normalize(name)
    if "description_cn" in df.columns:
        col = "description_cn"
    elif "description" in df.columns:
        col = "description"
    else:
        col = "Describrition" if "Describrition" in df.columns else "Description"
    series = df[col].astype(str).apply(_normalize)
    rows = df[series == name_norm]
    return [_compute_profit_for_price(row, price) for _, row in rows.iterrows()]


def _try_match_material_code(
    keywords: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
) -> Optional[dict[str, Any]]:
    code = (keywords or "").strip()
    if not _looks_like_material_code(code):
        return None
    return get_wanding_price_by_code(code, customer_level, price_library_path)


def match_fuzzy(
    keywords: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
    product_type: Optional[str] = None,
) -> Optional[dict[str, Any]]:
    """
    DataBase-style 模糊匹配，返回最佳单结果。
    返回 {code, matched_name, unit_price} 或 None。
    先按业务知识【字段匹配补充规则】扩展检索词，再做同义/外语替换（SYNONYM_GROUPS、QUERY_TERM_TO_CHINESE）。
    """
    keywords = (keywords or "").strip()
    hit = _try_match_material_code(keywords, customer_level, price_library_path)
    if hit:
        out: dict[str, Any] = {
            "code": (hit.get("code") or "").strip(),
            "matched_name": (hit.get("matched_name") or "")[:200],
            "unit_price": float(hit.get("unit_price", 0) or 0),
        }
        desc_en = str(hit.get("description_english") or "").strip()
        if desc_en:
            out["description_english"] = desc_en[:500]
            out["indonesian_name"] = desc_en[:500]
        supplier = str(hit.get("supplier") or "").strip()
        if supplier:
            out["supplier"] = supplier
        return out
    keywords = _apply_knowledge_expansion(keywords)
    keywords = _apply_pressure_expansion(keywords)
    keywords = _normalize_unicode_fractions(keywords)
    keywords = _normalize_keyword_terms(keywords)
    keywords = _apply_drat_thread_expansion(keywords)
    keywords = _strip_query_intent_terms(keywords)
    if not keywords:
        return None

    from inventory.config import config
    use_remote = price_library_path is None
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level, try_remote=use_remote)
    if df.empty:
        return None
    if product_type:
        if "Product_Type" not in df.columns:
            logger.warning("Product_Type 严格过滤失败：数据源缺少 Product_Type 列，filter=%s", product_type)
            return None
        df = df[df["Product_Type"].astype(str).str.strip() == product_type]
        if df.empty:
            logger.warning("Product_Type 过滤后无候选: %s", product_type)
            return None

    results = search_fuzzy(df, keywords)
    if not results:
        return None

    row_dict, _ = results[0]
    out: dict[str, Any] = {
        "code": (row_dict.get("code") or "").strip(),
        "matched_name": (row_dict.get("matched_name") or "")[:200],
        "unit_price": float(row_dict.get("unit_price", 0) or 0),
    }
    desc_en = str(row_dict.get("description_english") or "").strip()
    if desc_en:
        out["description_english"] = desc_en[:500]
        out["indonesian_name"] = desc_en[:500]
    supplier = str(row_dict.get("supplier") or "").strip()
    if supplier:
        out["supplier"] = supplier
    if use_remote:
        level = _normalize_price_level(customer_level)
        meta = get_remote_price_meta(level)
        if meta:
            out["price_source"] = meta.get("price_source")
            out["price_stale"] = meta.get("price_stale", False)
            out["price_stale_warning"] = meta.get("price_stale_warning")
            out["price_version_id"] = meta.get("price_version_id")
            out["price_version_number"] = meta.get("price_version_number")
    return out


def match_fuzzy_candidates(
    keywords: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
    max_candidates: int = 20,
    max_score_tiers: Optional[int] = None,
    min_score: Optional[float] = None,
    min_score_gap: Optional[float] = None,
    product_type: Optional[str] = None,
) -> List[dict[str, Any]]:
    """
    返回候选列表，每项含 code, matched_name, unit_price, score。
    - max_score_tiers 为 None：按分数排序取前 max_candidates 条。
    - max_score_tiers 为 N（如 2）：取分数前 N 档，每档全部返回（如 top1 有 3 条、top2 有 2 条则共 5 条）。
    - min_score：若最高分低于该阈值，则视为整体未命中，返回空列表。
    - min_score_gap：若 top1 分数与 top2 之差 ≥ 该值，则仅保留 top1 作为唯一高置信度候选。
    先按业务知识【字段匹配补充规则】扩展检索词，再做同义/外语替换。
    """
    keywords = (keywords or "").strip()
    hit = _try_match_material_code(keywords, customer_level, price_library_path)
    if hit:
        item: dict[str, Any] = {
            "code": hit["code"],
            "matched_name": hit["matched_name"],
            "unit_price": hit["unit_price"],
            "score": 1.0,
            "Product_Type": "",
        }
        desc_en = str(hit.get("description_english") or "").strip()
        if desc_en:
            item["description_english"] = desc_en
        supplier = str(hit.get("supplier") or "").strip()
        if supplier:
            item["supplier"] = supplier
        return [item]
    keywords = _apply_knowledge_expansion(keywords)
    keywords = _apply_pressure_expansion(keywords)
    keywords = _normalize_unicode_fractions(keywords)
    keywords = _normalize_keyword_terms(keywords)
    keywords = _apply_drat_thread_expansion(keywords)
    keywords = _strip_query_intent_terms(keywords)
    if not keywords:
        return []

    from inventory.config import config
    use_remote = price_library_path is None
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level, try_remote=use_remote)
    if df.empty:
        return []
    if product_type:
        if "Product_Type" not in df.columns:
            logger.warning("Product_Type 严格过滤失败：数据源缺少 Product_Type 列，filter=%s", product_type)
            return []
        df = df[df["Product_Type"].astype(str).str.strip() == product_type]
        if df.empty:
            logger.warning("Product_Type 过滤后无候选: %s", product_type)
            return []

    results = search_fuzzy(df, keywords)
    if not results:
        return []

    # 默认阈值从配置读取（如未显式传入）
    if min_score is None:
        min_score = getattr(config, "INVENTORY_MIN_SCORE", None)
    if min_score_gap is None:
        min_score_gap = getattr(config, "INVENTORY_MIN_SCORE_GAP", None)

    top_score = results[0][1]
    # 若最高分低于阈值，则视为整体未命中，直接返回空列表
    if isinstance(min_score, (int, float)) and top_score < float(min_score):
        return []

    # 若 top1 与 top2 分数差足够大，则视为唯一高置信度候选，直接截断为单候选
    if (
        isinstance(min_score_gap, (int, float))
        and len(results) >= 2
        and (top_score - results[1][1]) >= float(min_score_gap)
    ):
        results = results[:1]
    elif max_score_tiers is not None and max_score_tiers > 0:
        # 取前 max_score_tiers 个分数档，每档全部返回
        tiers: List[float] = []
        for _rd, score in results:
            if score not in tiers:
                tiers.append(score)
                if len(tiers) >= max_score_tiers:
                    break
        results = [(rd, s) for rd, s in results if s in tiers]
    else:
        results = results[:max_candidates]
    remote_meta: dict[str, Any] = {}
    if use_remote:
        level_key = _normalize_price_level(customer_level)
        remote_meta = get_remote_price_meta(level_key)
    out = []
    for row_dict, score in results:
        item: dict[str, Any] = {
            "code": (row_dict.get("code") or "").strip(),
            "matched_name": (row_dict.get("matched_name") or "")[:200],
            "unit_price": float(row_dict.get("unit_price", 0) or 0),
            "score": round(score, 4),
            "Product_Type": (row_dict.get("Product_Type") or "").strip(),
        }
        desc_en = str(row_dict.get("description_english") or "").strip()
        if desc_en:
            item["description_english"] = desc_en[:500]
        supplier = str(row_dict.get("supplier") or "").strip()
        if supplier:
            item["supplier"] = supplier
        if remote_meta:
            item["price_source"] = remote_meta.get("price_source")
            item["price_stale"] = remote_meta.get("price_stale", False)
            item["price_stale_warning"] = remote_meta.get("price_stale_warning")
            item["price_version_id"] = remote_meta.get("price_version_id")
            item["price_version_number"] = remote_meta.get("price_version_number")
        out.append(item)
    return out


def match_english_candidates(
    keywords: str,
    customer_level: str = "B",
    price_library_path: str | Path | None = None,
    max_candidates: int = 20,
    product_type: Optional[str] = None,
) -> List[dict[str, Any]]:
    """
    英文 query → Describrition_English CONTAINS 匹配。
    将 keywords 按空白/标点拆分：长度 ≥2 的片段，以及单独的数字规格（如 3、50）作为 token；
    每行需全部 token 均出现在英文描述（小写）中才收录。
    返回 [{code, matched_name, description_english, unit_price, source}, ...]，最多 max_candidates 条。
    matched_name 为中文 Describrition，保持下游结构一致。
    """
    from inventory.config import config

    use_remote = price_library_path is None
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level, try_remote=use_remote)
    if df.empty or "Describrition_English" not in df.columns:
        return []
    if product_type:
        if "Product_Type" not in df.columns:
            logger.warning("Product_Type 严格过滤失败：数据源缺少 Product_Type 列，filter=%s", product_type)
            return []
        df = df[df["Product_Type"].astype(str).str.strip() == product_type]
        if df.empty:
            logger.warning("Product_Type 过滤后无候选: %s", product_type)
            return []

    raw_tokens = re.split(r'[\s\"\'\-/\\]+', (keywords or "").strip())
    tokens: list[str] = []
    for t in raw_tokens:
        t = (t or "").strip()
        if not t:
            continue
        tl = t.lower()
        if len(tl) >= 2:
            tokens.append(tl)
        elif tl.isdigit():
            # 英寸/规格数字（如 3" pipe 中的 3）单独参与 CONTAINS
            tokens.append(tl)
    if not tokens:
        return []

    norm_kw = _normalize(keywords or "")
    results: list[dict[str, Any]] = []
    for row in df.itertuples(index=False):
        en_desc = str(getattr(row, "Describrition_English", "") or "").lower()
        if not en_desc:
            continue
        if all(t in en_desc for t in tokens):
            row_dict = {
                "code": str(getattr(row, "Material", "")).strip(),
                "matched_name": str(getattr(row, "Describrition", "")).strip(),
                "description_english": str(getattr(row, "Describrition_English", "")).strip(),
                "Product_Type": str(getattr(row, "Product_Type", "")).strip(),
                "supplier": str(getattr(row, "supplier", "") or getattr(row, "Supplier", "") or "").strip(),
                "unit_price": float(getattr(row, "unit_price", 0) or 0),
                "source": "英文字段匹配",
            }
            ok, bonus = _hard_filter_and_bonus(
                norm_kw,
                f"{row_dict['matched_name']} {row_dict['description_english']}",
                row_dict["code"],
                row_dict["Product_Type"],
            )
            if ok:
                row_dict["_rank_bonus"] = bonus
                results.append(row_dict)

    results.sort(key=lambda r: -float(r.get("_rank_bonus", 0.0)))
    for row in results:
        row.pop("_rank_bonus", None)
    return results[:max_candidates]


def get_wanding_price_by_code(
    code: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
) -> Optional[dict[str, Any]]:
    """
    按产品编号（Material）在万鼎价格表中精确查找，返回该 code 的单价及名称。
    用于历史匹配拿到 code 后，从万鼎表把价格补全。
    返回 {code, matched_name, unit_price, description_english} 或 None（万鼎表无此 code）。
    """
    code = (code or "").strip()
    if not code:
        return None
    from inventory.config import config
    use_remote = price_library_path is None
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level, try_remote=use_remote)
    if df.empty or "Material" not in df.columns:
        return None
    code_norm = _normalize_code_for_match(code)
    row = df[df["Material"].apply(lambda v: _normalize_code_for_match(v) == code_norm)]
    if row.empty:
        return None
    r = row.iloc[0]
    desc_en = str(r.get("Describrition_English", "") or r.get("description_english", "") or "").strip()
    result: dict[str, Any] = {
        "code": code,
        "matched_name": str(r.get("Describrition", "") or "")[:200],
        "unit_price": float(r.get("unit_price", 0) or 0),
    }
    if desc_en:
        result["description_english"] = desc_en[:500]
    supplier = str(r.get("supplier", "") or r.get("Supplier", "") or "").strip()
    if supplier:
        result["supplier"] = supplier
    if use_remote:
        level_key = _normalize_price_level(customer_level)
        meta = get_remote_price_meta(level_key)
        if meta:
            result["price_source"] = meta.get("price_source")
            result["price_stale"] = meta.get("price_stale", False)
            result["price_stale_warning"] = meta.get("price_stale_warning")
            result["price_version_id"] = meta.get("price_version_id")
            result["price_version_number"] = meta.get("price_version_number")
    return result
