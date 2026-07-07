#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Smoke: learn-by-data Section D — mapping pending + merge on VANTSING fixture."""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "python"))

import importlib.util

from inventory.services.mapping_table_matcher import load_mapping_df, search_mapping_fuzzy  # noqa: E402
from inventory.config import config  # noqa: E402
from quotation.layout import VANTSING_LAYOUT  # noqa: E402
from quotation.learn_by_data_mapping import (  # noqa: E402
    build_learn_by_data_mapping_row,
    check_learn_by_data_mapping_guards,
    is_section_d_eligible,
    load_mapping_pending_entries,
    mapping_search_text,
    section_d_trigger,
)
from quotation.mapping_pending_dispatch import handle_append_quotation_mapping_pending  # noqa: E402

FIXTURE = ROOT / "data" / "smoke" / "learn-by-data-vantsing-filled.xlsx"
MANIFEST = ROOT / "data" / "smoke" / "learn-by-data-vantsing-filled.manifest.json"
MERGE_PATH = ROOT / "python" / "scripts" / "merge_mapping_import.py"


def _load_merge_module():
    spec = importlib.util.spec_from_file_location("merge_mapping_import", MERGE_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _read_vantsing_rows(xlsx_path: Path) -> list[dict]:
    import openpyxl

    layout = VANTSING_LAYOUT
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for excel_row in range(layout.data_start_row, ws.max_row + 1):
        marker = ws.cell(excel_row, layout.total_col).value
        if marker and "Total Excluding PPN" in str(marker):
            break
        inquiry_name = str(ws.cell(excel_row, layout.inquiry_name_col).value or "").strip()
        inquiry_spec = str(ws.cell(excel_row, layout.inquiry_spec_col).value or "").strip()
        product_code = str(ws.cell(excel_row, layout.product_no_col).value or "").strip()
        quote_name = str(ws.cell(excel_row, layout.quote_name_col).value or "").strip()
        if not inquiry_name and not product_code:
            continue
        rows.append(
            {
                "excel_row": excel_row,
                "inquiry_name": inquiry_name,
                "inquiry_spec": inquiry_spec if inquiry_spec != "None" else "",
                "product_code": product_code,
                "quote_name": quote_name,
                "keywords": mapping_search_text(inquiry_name, inquiry_spec),
            }
        )
    wb.close()
    return rows


def _agent_pick_for_keywords(keywords: str) -> str:
    from inventory.services.match_and_inventory import match_quotation_union

    cands = match_quotation_union(keywords, customer_level="B")
    return str(cands[0]["code"]) if cands else ""


def run_smoke(*, fixture_path: Path | None = None, eval_manifest_path: Path | None = None) -> dict:
    fixture = fixture_path or FIXTURE
    eval_mode = eval_manifest_path is not None
    use_legacy_manifest = not eval_mode and fixture.resolve() == FIXTURE.resolve()
    scenario_by_row: dict[int, str] = {}
    agent_pick_by_row: dict[int, str] = {}
    expected_trigger_by_row: dict[int, str | None] = {}
    manifest_path = eval_manifest_path if eval_mode else (MANIFEST if use_legacy_manifest else None)
    if manifest_path and manifest_path.is_file():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        for item in manifest.get("rows", []):
            row_no = int(item["row"])
            scenario_by_row[row_no] = str(item.get("scenario") or "")
            if "agent_pick_for_eval" in item:
                agent_pick_by_row[row_no] = str(item.get("agent_pick_for_eval") or "")
            expected_trigger_by_row[row_no] = item.get("expected_section_d_trigger")

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        pending_path = tmp_path / "pending.jsonl"
        mapping_path = tmp_path / "mapping_table.xlsx"
        shutil.copy2(config.MAPPING_TABLE_PATH, mapping_path)
        try:
            mapping_path.chmod(0o666)
        except OSError:
            pass

        import quotation.learn_by_data_mapping as lbdm

        original_resolve = lbdm.resolve_mapping_pending_path
        lbdm.resolve_mapping_pending_path = lambda: pending_path

        import quotation.mapping_pending_dispatch as mpd

        mpd.load_mapping_pending_entries = lambda: load_mapping_pending_entries(pending_path)

        sheet_rows = _read_vantsing_rows(fixture)
        mismatch_targets = []
        gap_targets = []
        preview_ok = 0
        applied_ok = 0
        skipped = 0
        rejected = 0

        mapping_df = load_mapping_df(mapping_path)

        for row in sheet_rows:
            excel_row = row["excel_row"]
            scenario = scenario_by_row.get(excel_row, "")
            if use_legacy_manifest and scenario not in {"in-candidates", "not-in-candidates"}:
                continue
            if eval_mode and expected_trigger_by_row.get(excel_row) is None:
                skipped += 1
                continue
            agent_pick = agent_pick_by_row.get(excel_row) if excel_row in agent_pick_by_row else _agent_pick_for_keywords(row["keywords"])
            if not is_section_d_eligible(
                agent_pick_code=agent_pick,
                sheet_product_code=row["product_code"],
                inquiry_name=row["inquiry_name"],
                inquiry_spec=row["inquiry_spec"],
                mapping_df=mapping_df,
            ):
                skipped += 1
                continue

            trigger = section_d_trigger(
                agent_pick_code=agent_pick,
                sheet_product_code=row["product_code"],
                inquiry_name=row["inquiry_name"],
                inquiry_spec=row["inquiry_spec"],
                mapping_df=mapping_df,
            )

            guard = check_learn_by_data_mapping_guards(
                inquiry_name=row["inquiry_name"],
                inquiry_spec=row["inquiry_spec"],
                sheet_product_code=row["product_code"],
                source_file=fixture.name,
                source_sheet="Sheet1",
                source_row=excel_row,
                mapping_df=mapping_df,
                pending_entries=load_mapping_pending_entries(pending_path),
            )
            if guard["action"] in {"skip", "reject"}:
                if guard["action"] == "reject":
                    rejected += 1
                else:
                    skipped += 1
                continue

            preview = handle_append_quotation_mapping_pending(
                {
                    "inquiry_name": row["inquiry_name"],
                    "inquiry_spec": row["inquiry_spec"],
                    "product_code": row["product_code"],
                    "quotation_name": row["quote_name"],
                    "source_file": fixture.name,
                    "source_sheet": "Sheet1",
                    "source_row": excel_row,
                    "agent_pick_code": agent_pick,
                    "confirmed": False,
                }
            )
            if not preview.get("requires_confirmation"):
                continue
            preview_ok += 1

            apply_params = {
                "inquiry_name": row["inquiry_name"],
                "inquiry_spec": row["inquiry_spec"],
                "product_code": row["product_code"],
                "quotation_name": row["quote_name"],
                "source_file": fixture.name,
                "source_sheet": "Sheet1",
                "source_row": excel_row,
                "agent_pick_code": agent_pick,
                "confirmed": True,
            }
            if guard["action"] == "confirm_overwrite":
                apply_params["allow_overwrite"] = True
            applied = handle_append_quotation_mapping_pending(apply_params)
            if applied.get("applied"):
                applied_ok += 1
                if trigger == "gap":
                    gap_targets.append(row)
                else:
                    mismatch_targets.append(row)

        merge_mapping_import = _load_merge_module()
        merge_result = merge_mapping_import.merge_pending(
            mapping_path=mapping_path,
            pending_path=pending_path,
            dry_run=False,
        )

        lbdm.resolve_mapping_pending_path = original_resolve

        recall_hits = 0
        from inventory.services.mapping_table_matcher import invalidate_mapping_cache

        invalidate_mapping_cache()
        merged_df = load_mapping_df(mapping_path)
        recall_targets = mismatch_targets + gap_targets
        for row in recall_targets[:3]:
            hits = search_mapping_fuzzy(merged_df, row["keywords"])
            codes = {str(h[0].get("code") or "") for h in hits}
            if row["product_code"] in codes:
                recall_hits += 1

        result = {
            "fixture": str(fixture),
            "eval_mode": eval_mode,
            "mismatch_preview_ok": preview_ok,
            "mismatch_applied_ok": applied_ok,
            "gap_applied_n": len(gap_targets),
            "mismatch_applied_n": len(mismatch_targets),
            "skipped_or_rejected": skipped + rejected,
            "merge": merge_result,
            "recall_sample_hits": recall_hits,
            "recall_sample_n": min(3, len(recall_targets)),
            "pass": preview_ok > 0 and applied_ok > 0 and merge_result.get("appended", 0) > 0,
        }
        if eval_mode:
            result["pass"] = (
                result["pass"]
                and len(gap_targets) >= 1
                and len(mismatch_targets) >= 1
                and skipped + rejected >= 2
            )
        return result


def main() -> int:
    fixture_arg: Path | None = None
    eval_manifest_arg: Path | None = None
    args = sys.argv[1:]
    idx = 0
    while idx < len(args):
        token = args[idx]
        if token == "--eval-manifest" and idx + 1 < len(args):
            eval_manifest_arg = Path(args[idx + 1])
            idx += 2
            continue
        fixture_arg = Path(token)
        idx += 1

    result = run_smoke(fixture_path=fixture_arg, eval_manifest_path=eval_manifest_arg)
    out_name = "section-d-eval-smoke.json" if eval_manifest_arg else "section-d-smoke.json"
    out_path = ROOT / ".trellis" / "tasks" / "07-06-learn-by-data-price-library-enrich" / "test-records" / out_name
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["pass"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
