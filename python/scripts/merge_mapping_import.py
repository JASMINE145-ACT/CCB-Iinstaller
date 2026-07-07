#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Merge learn-by-data mapping_import_pending.jsonl into mapping_table.xlsx."""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

from inventory.config import config  # noqa: E402
from inventory.services.mapping_table_matcher import invalidate_mapping_cache  # noqa: E402
from quotation.learn_by_data_mapping import (  # noqa: E402
    load_mapping_pending_entries,
    normalized_mapping_key,
    resolve_mapping_pending_path,
)


def _load_xlsx_rows(path: Path) -> list[list]:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb.active or wb[wb.sheetnames[0]]
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _write_xlsx_rows(path: Path, rows: list[list]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp_path)
    wb.close()
    if path.exists():
        path.unlink()
    tmp_path.replace(path)


def _existing_keys(rows: list[list]) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    for row in rows[1:]:
        if len(row) < 3:
            continue
        name = str(row[0] or "").strip()
        spec = str(row[1] or "").strip() if len(row) > 1 else ""
        code = str(row[2] or "").strip()
        if not code:
            continue
        keys.add((normalized_mapping_key(name, spec), code))
    return keys


def _remove_rows_with_norm_key(rows: list[list], norm_key: str) -> tuple[list[list], int]:
    if not rows:
        return rows, 0
    header, *data = rows
    kept: list[list] = []
    removed = 0
    for row in data:
        if len(row) < 3:
            kept.append(row)
            continue
        name = str(row[0] or "").strip()
        spec = str(row[1] or "").strip() if len(row) > 1 else ""
        if normalized_mapping_key(name, spec) == norm_key:
            removed += 1
            continue
        kept.append(row)
    return [header, *kept], removed


def merge_pending(
    *,
    mapping_path: Path,
    pending_path: Path,
    dry_run: bool = False,
) -> dict:
    if not mapping_path.is_file():
        raise FileNotFoundError(f"mapping table not found: {mapping_path}")

    rows = _load_xlsx_rows(mapping_path)
    if not rows:
        rows = [["询价货物名称", "询价规格型号", "产品编号", "报价名称"]]

    existing = _existing_keys(rows)
    pending = load_mapping_pending_entries(pending_path)
    appended = 0
    skipped = 0
    overwritten = 0

    for entry in pending:
        if entry.get("status") == "merged":
            continue
        row = entry.get("row") if isinstance(entry.get("row"), dict) else entry
        if not isinstance(row, dict):
            continue
        name = str(row.get("inquiry_name") or "").strip()
        spec = str(row.get("inquiry_spec") or "").strip()
        code = str(row.get("product_code") or "").strip()
        quote_name = str(row.get("quotation_name") or "").strip()
        if not code:
            skipped += 1
            continue
        norm_key = normalized_mapping_key(name, spec)
        key = (norm_key, code)
        if key in existing:
            entry["status"] = "merged"
            skipped += 1
            continue
        if not dry_run:
            if entry.get("allow_overwrite"):
                rows, removed = _remove_rows_with_norm_key(rows, norm_key)
                overwritten += removed
                existing = _existing_keys(rows)
            rows.append([name, spec, code, quote_name])
            existing.add(key)
            entry["status"] = "merged"
            appended += 1
        else:
            appended += 1

    if not dry_run:
        _write_xlsx_rows(mapping_path, rows)
        updated_lines = [json.dumps(e, ensure_ascii=False) for e in pending]
        pending_path.write_text("\n".join(updated_lines) + ("\n" if updated_lines else ""), encoding="utf-8")
        invalidate_mapping_cache()

    return {
        "mapping_path": str(mapping_path),
        "pending_path": str(pending_path),
        "appended": appended,
        "skipped": skipped,
        "overwritten": overwritten,
        "dry_run": dry_run,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge mapping_import_pending.jsonl into mapping_table.xlsx")
    parser.add_argument("--mapping-path", default=config.MAPPING_TABLE_PATH)
    parser.add_argument("--pending-path", default=str(resolve_mapping_pending_path()))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    result = merge_pending(
        mapping_path=Path(args.mapping_path),
        pending_path=Path(args.pending_path),
        dry_run=args.dry_run,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
