#!/usr/bin/env python3
"""Prepare price-library xlsx for AionCore import (migration column mapping).

Fills empty price_a..price_d from tax columns without overwriting existing tier
prices. Rows with no quotable price after mapping are kept (catalog-only SKUs,
e.g. CP-ML* ceiling materials) so they appear in the org price library.
Duplicate material codes are collapsed by keeping the row with
`is_preferred_price=True` when exactly one exists (752 PE/LESSO overlap groups in
the 2026-05-15 workbook). Aligns with `data/data.Md` and the wanding matcher.
If multiple preferred rows exist, the first preferred row wins; if none, the last
row in workbook order wins.

Mapping (only when target cell is empty):
  price_b <- local_inc_tax, else local_exc_tax
  price_c <- factory_inc_tax, else factory_exc_tax

Usage:
  python scripts/org-phase0/prepare-price-library-import.py
  python scripts/org-phase0/prepare-price-library-import.py \\
    --input data/price_library_cleaned_2026_05_15.xlsx \\
    --output data/price_library_import_ready.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = (
    REPO_ROOT
    / ".trellis"
    / "tasks"
    / "06-30-quotation-supplier-remark"
    / "price_library_with_supplier_simple_draft.xlsx"
)
DEFAULT_OUTPUT = REPO_ROOT / "data" / "price_library_import_ready.xlsx"
DEFAULT_SKIPPED = REPO_ROOT / "data" / "price_library_import_skipped.json"
SHEET_NAME = "price_library"

PRICE_COLS = ("price_a", "price_b", "price_c", "price_d")
FALLBACK_FOR_B = ("local_inc_tax", "local_exc_tax")
FALLBACK_FOR_C = ("factory_inc_tax", "factory_exc_tax")

# Any of these satisfies import (matches AionCore excel.rs / data.Md).
QUOTABLE_PRICE_COLS = (
    "factory_inc_tax",
    "factory_exc_tax",
    "purchase_exc_tax",
    "price_a",
    "price_b",
    "price_c",
    "price_d",
    "price_d_low",
    "price_e",
    "local_exc_tax",
    "local_inc_tax",
    "rucika_pricelist_exc_vat11",
    "rucika_pricelist_inc_vat11",
    "rucika_quote_price_1",
    "rucika_quote_price_2",
    "pe_nominal_price",
    "pe_factory_price",
)


def col_index(header_row: tuple) -> dict[str, int]:
    return {str(h).strip().lower(): i for i, h in enumerate(header_row) if h}


def cell_value(row: tuple, idx: int | None):
    if idx is None:
        return None
    return row[idx]


def is_empty_price(value) -> bool:
    return value is None


def has_any_price(row: tuple, idx: dict[str, int]) -> bool:
    return any(not is_empty_price(cell_value(row, idx.get(c))) for c in QUOTABLE_PRICE_COLS if c in idx)


def is_preferred(value) -> bool:
    return value in (True, 1, "1", "true", "True", "TRUE")


def pick_dedupe_winner(
    candidates: list[tuple[int, tuple]],
    preferred_i: int | None,
) -> tuple[int, tuple, str]:
    """Return (source_row, row, reason) for one material's duplicate group."""
    if preferred_i is not None:
        preferred = [(sr, row) for sr, row in candidates if is_preferred(row[preferred_i])]
        if len(preferred) == 1:
            source_row, row = preferred[0]
            return source_row, row, "is_preferred_price"
        if len(preferred) > 1:
            source_row, row = preferred[0]
            return source_row, row, "multiple_preferred_kept_first"
    source_row, row = candidates[-1]
    return source_row, row, "kept_last_no_preferred"


def apply_fallback(row: list, idx: dict[str, int], target: str, sources: tuple[str, ...]) -> bool:
    target_i = idx.get(target)
    if target_i is None or not is_empty_price(row[target_i]):
        return False
    for source in sources:
        source_i = idx.get(source)
        if source_i is None:
            continue
        value = row[source_i]
        if not is_empty_price(value):
            row[target_i] = value
            return True
    return False


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--skipped", type=Path, default=DEFAULT_SKIPPED)
    parser.add_argument("--sheet", default=SHEET_NAME)
    args = parser.parse_args()

    if not args.input.is_file():
        print(f"input not found: {args.input}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(args.input)
    if args.sheet not in wb.sheetnames:
        print(f"sheet {args.sheet!r} not in workbook: {wb.sheetnames}", file=sys.stderr)
        return 1

    ws = wb[args.sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("workbook is empty", file=sys.stderr)
        return 1

    header = rows[0]
    idx = col_index(header)
    material_i = idx.get("material")
    if material_i is None:
        print("missing material column", file=sys.stderr)
        return 1

    if "description" not in idx and "description_cn" not in idx:
        print("missing description column", file=sys.stderr)
        return 1

    mapped_b = 0
    mapped_c = 0
    kept_unchanged = 0
    skipped: list[dict] = []
    no_price_catalog: list[dict] = []
    deduped: list[dict] = []
    by_material: dict[str, list[tuple[int, tuple]]] = {}
    preferred_i = idx.get("is_preferred_price")
    supplier_i = idx.get("supplier")

    for source_row, raw in enumerate(rows[1:], start=2):
        if raw is None:
            continue
        row = list(raw)
        material = row[material_i]
        if material is None or str(material).strip() == "":
            skipped.append(
                {
                    "source_row": source_row,
                    "material": "",
                    "description": "",
                    "reason": "empty material",
                }
            )
            continue

        material_code = str(material).strip()
        had_abcd = any(not is_empty_price(cell_value(row, idx.get(c))) for c in PRICE_COLS if c in idx)
        if apply_fallback(row, idx, "price_b", FALLBACK_FOR_B):
            mapped_b += 1
        if apply_fallback(row, idx, "price_c", FALLBACK_FOR_C):
            mapped_c += 1

        if had_abcd:
            kept_unchanged += 1

        if not has_any_price(row, idx):
            no_price_catalog.append(
                {
                    "source_row": source_row,
                    "material": material_code,
                    "description": str(row[idx["description"]]).strip()
                    if "description" in idx and row[idx["description"]] is not None
                    else "",
                    "reason": "catalog_only_no_quotable_price",
                }
            )

        by_material.setdefault(material_code, []).append((source_row, tuple(row)))

    out_rows: list[tuple] = [header]
    supplier_non_empty = 0
    supplier_by_material: dict[str, set[str]] = {}
    for material_code, candidates in by_material.items():
        winner_source, winner_tuple, reason = pick_dedupe_winner(candidates, preferred_i)
        out_rows.append(winner_tuple)
        if supplier_i is not None:
            supplier_val = winner_tuple[supplier_i]
            if supplier_val is not None and str(supplier_val).strip():
                supplier_non_empty += 1
                supplier_by_material.setdefault(material_code, set()).add(str(supplier_val).strip())
        if len(candidates) > 1:
            for source_row, _ in candidates:
                if source_row == winner_source:
                    continue
                deduped.append(
                    {
                        "source_row": source_row,
                        "material": material_code,
                        "kept_source_row": winner_source,
                        "reason": reason,
                    }
                )

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = args.sheet
    for row in out_rows:
        out_ws.append(row)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.skipped.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(args.output)

    manifest = {
        "input": str(args.input),
        "output": str(args.output),
        "sheet": args.sheet,
        "source_data_rows": len(rows) - 1,
        "importable_rows": len(out_rows) - 1,
        "skipped_rows": len(skipped),
        "no_price_catalog_rows": len(no_price_catalog),
        "mapped_price_b": mapped_b,
        "mapped_price_c": mapped_c,
        "rows_with_existing_abcd": kept_unchanged,
        "deduped_rows": len(deduped),
        "supplier_column_present": supplier_i is not None,
        "supplier_non_empty_count": supplier_non_empty,
        "multi_supplier_material_count": sum(1 for s in supplier_by_material.values() if len(s) > 1),
        "skipped": skipped,
        "no_price_catalog": no_price_catalog,
        "deduped": deduped,
    }
    args.skipped.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Wrote {args.output} ({manifest['importable_rows']} importable rows)")
    print(
        f"Skipped manifest: {args.skipped} "
        f"({manifest['skipped_rows']} skipped, {manifest['no_price_catalog_rows']} catalog-only, "
        f"{manifest['deduped_rows']} deduped)"
    )
    print(f"Mapped price_b: {mapped_b}, price_c: {mapped_c}")
    if supplier_i is not None:
        print(
            f"Supplier: {supplier_non_empty} non-empty rows, "
            f"{manifest['multi_supplier_material_count']} materials with multiple suppliers in source"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
