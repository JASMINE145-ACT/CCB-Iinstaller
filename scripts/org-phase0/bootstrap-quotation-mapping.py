#!/usr/bin/env python3
"""Bootstrap org quotation-mapping from mapping_table.xlsx (qmap_* tables only).

Uses org HTTP API — does not touch price_library, org_knowledge, or users tables.

Usage (Windows, reads scripts/org-phase0/env.local):
  python scripts/org-phase0/bootstrap-quotation-mapping.py --dry-run
  python scripts/org-phase0/bootstrap-quotation-mapping.py

Usage (VPS, reads /root/org-phase0.env):
  python3 /opt/aionorg/scripts/bootstrap-quotation-mapping.py \\
    --path /opt/aionorg/data/mapping_table.xlsx \\
    --org-url http://127.0.0.1:13401
"""
from __future__ import annotations

import argparse
import http.cookiejar
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "python"))

from admin.org_http_csrf import ORG_CSRF_HEADER_NAME, bootstrap_org_csrf, build_cookie_opener  # noqa: E402
from quotation.learn_by_data_mapping import normalized_mapping_key  # noqa: E402


def _load_env_file(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.is_file():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def _resolve_env() -> dict[str, str]:
    candidates = [
        Path(__file__).resolve().parent / "env.local",
        Path("/root/org-phase0.env"),
    ]
    merged: dict[str, str] = {}
    for path in candidates:
        merged.update(_load_env_file(path))
    return merged


def _unwrap_api(payload: dict[str, Any]) -> dict[str, Any]:
    if isinstance(payload.get("data"), dict):
        return payload["data"]
    return payload


class OrgMappingBootstrapClient:
    def __init__(self, base_url: str, token: str) -> None:
        self.base = base_url.rstrip("/")
        self.token = token
        self._jar = http.cookiejar.CookieJar()
        self._opener = build_cookie_opener(self._jar)
        self._csrf_token = ""

    def _ensure_csrf(self, *, force: bool = False) -> str:
        if force or not self._csrf_token:
            self._csrf_token = bootstrap_org_csrf(self.base, self._jar, opener_factory=lambda _: self._opener)
        return self._csrf_token

    def _request(self, method: str, path: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {self.token}",
        }
        data = None
        if payload is not None:
            headers["Content-Type"] = "application/json"
            headers[ORG_CSRF_HEADER_NAME] = self._ensure_csrf()
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(f"{self.base}{path}", data=data, headers=headers, method=method)
        try:
            with self._opener.open(req, timeout=60) as resp:
                raw = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if payload is not None and e.code == 403:
                headers[ORG_CSRF_HEADER_NAME] = self._ensure_csrf(force=True)
                req = urllib.request.Request(f"{self.base}{path}", data=data, headers=headers, method=method)
                with self._opener.open(req, timeout=60) as resp:
                    raw = json.loads(resp.read().decode("utf-8"))
            else:
                body = e.read().decode("utf-8", errors="replace")
                raise RuntimeError(f"HTTP {e.code} {method} {path}: {body[:500]}") from e
        if not isinstance(raw, dict):
            raise RuntimeError(f"Unexpected response for {path}")
        return _unwrap_api(raw)

    def get_active_rows(self) -> list[dict[str, Any]]:
        data = self._request("GET", "/api/quotation-mapping/active")
        rows = data.get("rows") or []
        return rows if isinstance(rows, list) else []

    def get_draft(self) -> dict[str, Any]:
        return self._request("GET", "/api/quotation-mapping/draft")

    def append_draft_item(self, fields: dict[str, Any]) -> dict[str, Any]:
        return self._request("POST", "/api/quotation-mapping/draft/items", fields)

    def publish_draft(self, *, reason: str, revision: int) -> dict[str, Any]:
        return self._request(
            "POST",
            "/api/quotation-mapping/draft/publish",
            {"reason": reason, "revision": revision},
        )


def org_login(base_url: str, username: str, password: str) -> str:
    body = json.dumps({"username": username, "password": password}).encode("utf-8")
    req = urllib.request.Request(
        f"{base_url.rstrip('/')}/login",
        data=body,
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    token = str(payload.get("token") or "").strip()
    if not payload.get("success") or not token:
        raise RuntimeError(f"org login failed: {payload.get('message') or payload}")
    return token


def load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(max_col=4, min_row=2), start=2):
        cells = [getattr(c, "value", None) for c in row]
        if len(cells) < 4:
            continue
        inquiry_name = str(cells[0] or "").strip()
        inquiry_spec = str(cells[1] or "").strip() if cells[1] is not None else ""
        code_raw = cells[2]
        if code_raw is None:
            continue
        product_code = str(code_raw).strip()
        quotation_name = str(cells[3] or "").strip()
        if not product_code:
            continue
        if not inquiry_name and not inquiry_spec:
            continue
        rows.append(
            {
                "inquiry_name": inquiry_name,
                "inquiry_spec": inquiry_spec,
                "product_code": product_code,
                "quotation_name": quotation_name,
                "source_file": path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "norm_key": normalized_mapping_key(inquiry_name, inquiry_spec),
                "allow_overwrite": True,
            }
        )
    wb.close()
    return rows


def dedupe_by_norm_key(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = str(row.get("norm_key") or "").strip()
        if key:
            by_key[key] = row
    return list(by_key.values())


def resolve_default_path() -> Path:
    vps_path = Path("/opt/aionorg/data/mapping_table.xlsx")
    if vps_path.is_file():
        return vps_path
    return REPO_ROOT / "data" / "mapping_table.xlsx"


def main() -> int:
    parser = argparse.ArgumentParser(description="Bootstrap org quotation-mapping from xlsx")
    parser.add_argument("--path", type=Path, default=None, help="mapping_table.xlsx path")
    parser.add_argument("--org-url", default="", help="Override ORG_CENTER_URL / ORG_SERVER_URL")
    parser.add_argument("--dry-run", action="store_true", help="Parse xlsx and print stats only")
    parser.add_argument(
        "--skip-if-active",
        type=int,
        default=1,
        metavar="N",
        help="Abort when active row count >= N (default: 1)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Import/publish even when active already has rows",
    )
    parser.add_argument(
        "--reason",
        default="bootstrap from mapping_table.xlsx",
        help="Publish reason string",
    )
    args = parser.parse_args()

    env = _resolve_env()
    base_url = (args.org_url or env.get("ORG_CENTER_URL") or env.get("ORG_SERVER_URL") or "").strip()
    username = (env.get("ORG_ADMIN_USER") or "admin").strip()
    password = (env.get("ORG_ADMIN_PASSWORD") or "").strip()

    xlsx_path = args.path or resolve_default_path()
    if not xlsx_path.is_file():
        print(f"ERROR: xlsx not found: {xlsx_path}", file=sys.stderr)
        return 1

    raw_rows = load_xlsx_rows(xlsx_path)
    rows = dedupe_by_norm_key(raw_rows)
    print(f"xlsx: {xlsx_path} ({xlsx_path.stat().st_size} bytes)", flush=True)
    print(f"rows read: {len(raw_rows)} | unique norm_key: {len(rows)}", flush=True)

    if args.dry_run:
        if rows:
            sample = rows[0]
            print("sample:", json.dumps(sample, ensure_ascii=False))
        return 0

    if not base_url:
        print("ERROR: set ORG_CENTER_URL in env.local or pass --org-url", file=sys.stderr)
        return 1
    if not password:
        print("ERROR: ORG_ADMIN_PASSWORD missing in env.local or /root/org-phase0.env", file=sys.stderr)
        return 1

    print(f"org: {base_url} user={username}", flush=True)
    token = org_login(base_url, username, password)
    client = OrgMappingBootstrapClient(base_url, token)
    client._ensure_csrf()

    active_rows = client.get_active_rows()
    print(f"active rows before: {len(active_rows)}", flush=True)
    if not args.force and len(active_rows) >= args.skip_if_active:
        print(
            f"SKIP: active already has {len(active_rows)} rows (use --force to merge/publish anyway)"
        )
        return 0

    for index, row in enumerate(rows, start=1):
        client.append_draft_item(row)
        if index == 1 or index % 100 == 0 or index == len(rows):
            print(f"  draft append {index}/{len(rows)}", flush=True)

    draft = client.get_draft()
    revision = int(draft.get("revision") or 0)
    item_count = len(draft.get("items") or [])
    print(f"draft revision={revision} items={item_count}")

    if item_count == 0:
        print("ERROR: draft has no items after append", file=sys.stderr)
        return 1

    version = client.publish_draft(reason=args.reason, revision=revision)
    print(
        "published:",
        json.dumps(
            {
                "version_number": version.get("version_number"),
                "item_count": version.get("item_count"),
                "reason": version.get("reason"),
            },
            ensure_ascii=False,
        ),
    )

    active_after = client.get_active_rows()
    print(f"active rows after: {len(active_after)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
